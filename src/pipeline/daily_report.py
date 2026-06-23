"""Daily Top10 report generation for live predictions."""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

import pandas as pd

from src.pipeline.config import DailyUpdateConfig


TOP10_COLUMNS: tuple[str, ...] = (
    "prediction_date",
    "AI Rank",
    "AI Percentile",
    "AI Score",
    "ticker",
    "ticker_name",
    "Expected Return(%)",
    "Gap Forecast(%)",
    "Intraday Forecast(%)",
    "prev_close",
    "pred_open_price",
    "pred_close_price",
)

FORBIDDEN_LIVE_COLUMNS: set[str] = {
    "Actual Return(%)",
    "Prediction Error(%)",
    "Hit",
    "Hit Binary",
    "Prediction Quality",
}


@dataclass(frozen=True)
class DailyReportResult:
    """Daily Top10 report output summary."""

    top10_report_csv: str
    top10_report_xlsx: str
    daily_summary_report: str
    top10_tickers: list[str]
    top10_ticker_names: list[str]
    top10_average_ai_score: float
    top10_average_expected_return: float
    warnings: list[str] = field(default_factory=list)


def generate_daily_top10_report(
    config: DailyUpdateConfig,
    prediction_date: str | pd.Timestamp,
    prediction_csv: str | Path | None = None,
    prediction_parquet: str | Path | None = None,
    old_data_warning: str | None = None,
    us_market_holiday_detected: bool = False,
    holiday_date: str | None = None,
    sources_using_prior_trading_day: list[str] | None = None,
) -> DailyReportResult:
    """Generate daily Top10 CSV, Excel, and markdown summary reports."""
    predictions = load_daily_predictions(config, prediction_date, prediction_csv, prediction_parquet)
    enriched, warnings = enrich_predictions_for_report(predictions)
    top10 = enriched.head(10).loc[:, TOP10_COLUMNS].copy()
    output_paths = report_paths(config, prediction_date)
    output_paths["dir"].mkdir(parents=True, exist_ok=True)
    top10.to_csv(output_paths["csv"], index=False, encoding="utf-8-sig")
    write_top10_excel(top10, output_paths["xlsx"])
    if old_data_warning:
        warnings.append(old_data_warning)
    write_daily_summary(
        enriched,
        top10,
        output_paths["summary"],
        warnings,
        us_market_holiday_detected=us_market_holiday_detected,
        holiday_date=holiday_date,
        sources_using_prior_trading_day=sources_using_prior_trading_day,
    )
    return DailyReportResult(
        top10_report_csv=str(output_paths["csv"]),
        top10_report_xlsx=str(output_paths["xlsx"]),
        daily_summary_report=str(output_paths["summary"]),
        top10_tickers=top10["ticker"].astype(str).tolist(),
        top10_ticker_names=top10["ticker_name"].astype(str).tolist(),
        top10_average_ai_score=float(top10["AI Score"].mean()) if not top10.empty else 0.0,
        top10_average_expected_return=float(predictions.nlargest(10, "ranking_score")["expected_return"].mean())
        if "expected_return" in predictions.columns and not predictions.empty
        else 0.0,
        warnings=warnings,
    )


def load_daily_predictions(
    config: DailyUpdateConfig,
    prediction_date: str | pd.Timestamp,
    prediction_csv: str | Path | None = None,
    prediction_parquet: str | Path | None = None,
) -> pd.DataFrame:
    """Load daily predictions from explicit paths or configured default paths."""
    compact = pd.Timestamp(prediction_date).strftime("%Y%m%d")
    parquet_path = Path(prediction_parquet) if prediction_parquet else config.resolve_path("daily_prediction_dir") / f"predictions_{compact}.parquet"
    csv_path = Path(prediction_csv) if prediction_csv else config.resolve_path("daily_prediction_dir") / f"predictions_{compact}.csv"
    if parquet_path.exists():
        data = pd.read_parquet(parquet_path)
    elif csv_path.exists():
        data = pd.read_csv(csv_path, dtype={"ticker": str})
    else:
        raise FileNotFoundError(f"No daily prediction file found for {compact}")
    data["ticker"] = data["ticker"].astype(str).str.zfill(6)
    if "ticker_name" not in data.columns:
        data["ticker_name"] = "UNKNOWN"
    return data


def enrich_predictions_for_report(predictions: pd.DataFrame) -> tuple[pd.DataFrame, list[str]]:
    """Add AI score, AI rank, AI percentile, and display-formatted columns."""
    data = predictions.copy()
    warnings: list[str] = []
    data = data.sort_values("ranking_score", ascending=False).reset_index(drop=True)
    universe_count = len(data)
    if universe_count == 0:
        raise ValueError("Daily prediction file has no rows")

    ranks = pd.Series(range(1, universe_count + 1), index=data.index)
    data["AI Rank"] = ranks.astype(str) + " / " + str(universe_count)
    percentile = (ranks / universe_count * 100).round().clip(lower=1, upper=100).astype(int)
    data["AI Percentile"] = "Top " + percentile.astype(str) + "%"
    min_score = float(data["ranking_score"].min())
    max_score = float(data["ranking_score"].max())
    if max_score == min_score:
        data["AI Score"] = 50.0
        warnings.append("ranking_score_min_equals_max_ai_score_set_to_50")
    else:
        data["AI Score"] = ((data["ranking_score"] - min_score) / (max_score - min_score) * 100).round(1)

    data["Expected Return(%)"] = data["expected_return"].map(format_percent)
    data["Gap Forecast(%)"] = data["pred_gap"].map(format_percent)
    data["Intraday Forecast(%)"] = data["pred_intraday"].map(format_percent)
    for column in ["prev_close", "pred_open_price", "pred_close_price"]:
        data[column] = data[column].round().astype("int64")
    forbidden = FORBIDDEN_LIVE_COLUMNS & set(data.columns)
    if forbidden:
        data = data.drop(columns=sorted(forbidden))
    return data, warnings


def format_percent(value: float) -> str:
    """Format decimal return as a percentage string."""
    return f"{float(value) * 100:.2f}%"


def report_paths(config: DailyUpdateConfig, prediction_date: str | pd.Timestamp) -> dict[str, Path]:
    """Return daily report output paths."""
    compact = pd.Timestamp(prediction_date).strftime("%Y%m%d")
    report_dir = config.resolve_path("daily_report_dir")
    return {
        "dir": report_dir,
        "csv": report_dir / f"top10_{compact}.csv",
        "xlsx": report_dir / f"top10_{compact}.xlsx",
        "summary": report_dir / f"daily_update_summary_{compact}.md",
    }


def write_feature_source_failure_summary(
    config: DailyUpdateConfig,
    report_date: str | pd.Timestamp,
    expected_feature_date: str,
    failed_sources: list[str],
    actual_source_dates: dict[str, str | None],
    target_update_date: str | None = None,
    latest_clean_data_date: str | None = None,
    stale_data_blocked: bool = False,
) -> Path:
    """Write a production-stop summary when feature sources are incomplete."""
    output_paths = report_paths(config, report_date)
    output_paths["dir"].mkdir(parents=True, exist_ok=True)
    target_date = target_update_date or expected_feature_date
    lines = [
        "# Daily Update Summary",
        "",
        "Status: FAILED",
        "Reason: Feature Source Completeness Failed",
        f"Expected Feature Date: {expected_feature_date}",
    ]
    if stale_data_blocked:
        lines.append("Production stopped because latest available feature source data is stale.")
    lines.extend(
        [
            f"Target Update Date: {target_date}",
            f"Latest Clean Data Date: {latest_clean_data_date or 'UNKNOWN'}",
            "",
            "Actual source dates:",
        ]
    )
    for source in ("krx", "nasdaq", "sp500", "vix", "wti", "usdkrw"):
        actual = actual_source_dates.get(source) or "missing"
        lines.append(f"- {display_source_name(source)}: {actual}")
    lines.extend(["", "Failed sources:"])
    for source in failed_sources:
        actual = actual_source_dates.get(source) or "missing"
        lines.append(f"- {display_source_name(source)}: expected {expected_feature_date}, actual {actual}")
    lines.extend(
        [
            "",
            "Top10: NOT GENERATED",
            "Prediction: NOT EXECUTED",
            "Retraining: NOT EXECUTED",
            "Feature Update: NOT EXECUTED",
            "",
        ]
    )
    output_paths["summary"].write_text("\n".join(lines), encoding="utf-8")
    return output_paths["summary"]


def display_source_name(source: str) -> str:
    """Return display name for a feature source."""
    display_names = {
        "krx": "KRX",
        "nasdaq": "NASDAQ",
        "sp500": "S&P500",
        "vix": "VIX",
        "wti": "WTI",
        "usdkrw": "USD/KRW",
        "us10y": "US10Y",
        "gold": "Gold",
        "dxy": "DXY",
    }
    return display_names.get(source, source.upper())


def write_top10_excel(top10: pd.DataFrame, path: Path) -> None:
    """Write Top10 Excel report with usability formatting."""
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        top10.to_excel(writer, index=False, sheet_name="Top10")
        explanation_rows = [
            ("AI Score", "0-100 normalized ranking score; higher means stronger model preference; not expected return."),
            ("AI Percentile", "Top x% within today's prediction universe."),
            ("AI Rank", "Rank among all predicted stocks."),
            ("Expected Return(%)", "pred_gap + pred_intraday."),
            ("Gap Forecast(%)", "Predicted previous close to open return."),
            ("Intraday Forecast(%)", "Predicted open to close return."),
            ("Price columns", "KRW integer display."),
        ]
        pd.DataFrame(explanation_rows, columns=["Term", "Meaning"]).to_excel(
            writer,
            index=False,
            sheet_name="Explanation",
        )
        workbook = writer.book
        worksheet = writer.sheets["Top10"]
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        for column_cells in worksheet.columns:
            max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 10), 28)
        price_columns = {"prev_close", "pred_open_price", "pred_close_price"}
        for idx, column_name in enumerate(top10.columns, start=1):
            letter = worksheet.cell(row=1, column=idx).column_letter
            if column_name in price_columns:
                for row in range(2, len(top10) + 2):
                    worksheet[f"{letter}{row}"].number_format = "#,##0"
            if column_name == "AI Score":
                for row in range(2, len(top10) + 2):
                    worksheet[f"{letter}{row}"].number_format = "0.0"
                add_ai_score_formatting(workbook, worksheet, letter, len(top10))
            if column_name == "Expected Return(%)":
                add_return_formatting(workbook, worksheet, letter, len(top10))
        explanation = writer.sheets["Explanation"]
        explanation.freeze_panes = "A2"
        explanation.auto_filter.ref = explanation.dimensions
        explanation.column_dimensions["A"].width = 24
        explanation.column_dimensions["B"].width = 90


def add_ai_score_formatting(workbook, worksheet, letter: str, row_count: int) -> None:
    """Apply AI score color bands."""
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import PatternFill

    cell_range = f"{letter}2:{letter}{row_count + 1}"
    bands = [
        ("between", ["90", "100"], "006100"),
        ("between", ["80", "89.999"], "C6EFCE"),
        ("between", ["70", "79.999"], "FFEB9C"),
        ("between", ["60", "69.999"], "F4B183"),
        ("lessThan", ["60"], "FFC7CE"),
    ]
    for operator, formulas, color in bands:
        worksheet.conditional_formatting.add(
            cell_range,
            CellIsRule(operator=operator, formula=formulas, fill=PatternFill("solid", fgColor=color)),
        )


def add_return_formatting(workbook, worksheet, letter: str, row_count: int) -> None:
    """Apply positive/negative color hints to expected return text cells."""
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.styles import Font

    cell_range = f"{letter}2:{letter}{row_count + 1}"
    worksheet.conditional_formatting.add(
        cell_range,
        FormulaRule(formula=[f'VALUE(SUBSTITUTE({letter}2,"%",""))>0'], font=Font(color="0000FF")),
    )
    worksheet.conditional_formatting.add(
        cell_range,
        FormulaRule(formula=[f'VALUE(SUBSTITUTE({letter}2,"%",""))<0'], font=Font(color="FF0000")),
    )


def write_daily_summary(
    enriched: pd.DataFrame,
    top10: pd.DataFrame,
    path: Path,
    warnings: list[str],
    us_market_holiday_detected: bool = False,
    holiday_date: str | None = None,
    sources_using_prior_trading_day: list[str] | None = None,
) -> None:
    """Write markdown summary for the daily update."""
    lines = [
        "# Daily Update Summary",
        "",
        f"- Prediction rows: {len(enriched)}",
        f"- Top10 rows: {len(top10)}",
        f"- Prediction date: {enriched['prediction_date'].iloc[0] if 'prediction_date' in enriched.columns else 'UNKNOWN'}",
        f"- Unique tickers: {enriched['ticker'].nunique()}",
        f"- Top10 average AI Score: {top10['AI Score'].mean():.2f}",
        f"- Top10 average Expected Return: {top10['Expected Return(%)'].map(lambda x: float(str(x).rstrip('%'))).mean():.2f}%",
        "",
        "## Top10",
        "",
    ]
    for _, row in top10.iterrows():
        lines.append(f"- {row['ticker']} {row['ticker_name']}: AI Score {row['AI Score']:.1f}")
    if us_market_holiday_detected:
        source_names = [display_source_name(source) for source in (sources_using_prior_trading_day or [])]
        lines.extend(
            [
                "",
                "## US Market Holiday Detected",
                "",
                f"- Holiday Date: {holiday_date or 'UNKNOWN'}",
                f"- Sources Using Prior Trading Day: {', '.join(source_names) or 'UNKNOWN'}",
            ]
        )
    if warnings:
        lines.extend(["", "## Warnings", ""])
        lines.extend(f"- {warning}" for warning in warnings)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
