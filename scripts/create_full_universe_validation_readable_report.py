"""Create readable reports for full-universe validation predictions."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[1]
PREDICTIONS_CSV = PROJECT_ROOT / "outputs" / "walk_forward_full_universe_rolling" / "predictions.csv"
TICKER_NAMES_CSV = PROJECT_ROOT / "data" / "metadata" / "ticker_names.csv"
REPORT_DIR = PROJECT_ROOT / "reports"

READABLE_CSV = REPORT_DIR / "full_universe_validation_predictions_readable.csv"
READABLE_XLSX = REPORT_DIR / "full_universe_validation_predictions_readable.xlsx"
TOP10_CSV = REPORT_DIR / "full_universe_validation_top10_by_date.csv"
TOP10_XLSX = REPORT_DIR / "full_universe_validation_top10_by_date.xlsx"
SUMMARY_MD = REPORT_DIR / "full_universe_validation_predictions_summary.md"

SAMSUNG_NAME = "\uc0bc\uc131\uc804\uc790"

PERCENT_COLUMNS = {
    "expected_return": "Expected Return(%)",
    "pred_gap": "Gap Forecast(%)",
    "pred_intraday": "Intraday Forecast(%)",
    "target_gap": "Actual Gap(%)",
    "target_intraday": "Actual Intraday(%)",
}
PRICE_COLUMNS = ["prev_close", "pred_open_price", "pred_close_price"]

READABLE_ORDER = [
    "date",
    "AI Rank",
    "AI Percentile",
    "AI Score",
    "ticker",
    "ticker_name",
    "Expected Return(%)",
    "Actual Return(%)",
    "Prediction Error(%)",
    "Hit",
    "Hit Binary",
    "Prediction Quality",
    "Gap Forecast(%)",
    "Intraday Forecast(%)",
    "Actual Gap(%)",
    "Actual Intraday(%)",
    "prev_close",
    "pred_open_price",
    "pred_close_price",
    "ranking_score",
    "expected_return",
    "pred_gap",
    "pred_intraday",
    "target_gap",
    "target_intraday",
]

REMAINING_PRIORITY = [
    "target_ranking",
    "model_version",
    "train_start_date",
    "train_end_date",
    "prediction_date",
]

TOP10_COLUMNS = [
    "date",
    "rank",
    "ticker",
    "ticker_name",
    "AI Score",
    "AI Percentile",
    "AI Rank",
    "Expected Return(%)",
    "Actual Return(%)",
    "Prediction Error(%)",
    "Hit",
    "Hit Binary",
    "Prediction Quality",
    "Gap Forecast(%)",
    "Intraday Forecast(%)",
    "Actual Gap(%)",
    "Actual Intraday(%)",
    "prev_close",
    "pred_open_price",
    "pred_close_price",
]


def load_predictions() -> pd.DataFrame:
    """Load full-universe predictions without modifying the source file."""
    if not PREDICTIONS_CSV.exists():
        raise FileNotFoundError(f"Missing predictions file: {PREDICTIONS_CSV}")
    df = pd.read_csv(PREDICTIONS_CSV, dtype={"ticker": str}, encoding="utf-8-sig")
    df["ticker"] = df["ticker"].astype(str).str.zfill(6)
    for column in ["date", "train_start_date", "train_end_date", "prediction_date"]:
        if column in df.columns:
            df[column] = pd.to_datetime(df[column])
    return df


def load_ticker_names() -> pd.DataFrame:
    """Load ticker name metadata."""
    if not TICKER_NAMES_CSV.exists():
        raise FileNotFoundError(f"Missing ticker mapping file: {TICKER_NAMES_CSV}")
    mapping = pd.read_csv(TICKER_NAMES_CSV, dtype={"ticker": str}, encoding="utf-8-sig")
    required = {"ticker", "ticker_name"}
    missing = required - set(mapping.columns)
    if missing:
        raise ValueError(f"ticker_names.csv missing columns: {sorted(missing)}")
    mapping["ticker"] = mapping["ticker"].astype(str).str.zfill(6)
    mapping["ticker_name"] = mapping["ticker_name"].fillna("UNKNOWN").astype(str)
    return mapping.loc[:, ["ticker", "ticker_name"]].drop_duplicates("ticker")


def ai_score(ranking_score: pd.Series) -> pd.Series:
    """Normalize ranking_score to 0-100 over the full prediction dataset."""
    min_score = ranking_score.min()
    max_score = ranking_score.max()
    if pd.isna(min_score) or pd.isna(max_score) or max_score == min_score:
        return pd.Series(50.0, index=ranking_score.index)
    return (100 * (ranking_score - min_score) / (max_score - min_score)).round(1)


def format_percent(value: object) -> str:
    """Format decimal value as percentage with two decimals."""
    if pd.isna(value):
        return ""
    return f"{float(value) * 100:.2f}%"


def create_readable(predictions: pd.DataFrame, ticker_names: pd.DataFrame) -> pd.DataFrame:
    """Create sorted, human-readable predictions with AI score fields."""
    base = predictions.drop(columns=["ticker_name"], errors="ignore").merge(
        ticker_names,
        on="ticker",
        how="left",
        validate="many_to_one",
    )
    base["ticker_name"] = base["ticker_name"].fillna("UNKNOWN")
    base = base.sort_values(["date", "ranking_score"], ascending=[True, False]).reset_index(drop=True)

    base["AI Score"] = ai_score(base["ranking_score"])
    rank = base.groupby("date")["ranking_score"].rank(method="first", ascending=False).astype(int)
    count = base.groupby("date")["ticker"].transform("count").astype(int)
    percentile = (rank / count * 100).round().clip(lower=1).astype(int)
    base["AI Percentile"] = percentile.map(lambda value: f"Top {value}%")
    base["AI Rank"] = rank.astype(str) + " / " + count.astype(str)

    for source, display_name in PERCENT_COLUMNS.items():
        base[display_name] = base[source].map(format_percent)
    add_evaluation_columns(base)
    for column in PRICE_COLUMNS:
        if column in base.columns:
            base[column] = base[column].round().astype("Int64")

    ordered = READABLE_ORDER + [column for column in REMAINING_PRIORITY if column in base.columns]
    remaining = [column for column in base.columns if column not in ordered]
    return base.loc[:, ordered + remaining].copy()


def create_top10(readable: pd.DataFrame) -> pd.DataFrame:
    """Create daily Top10 report."""
    top10 = readable.copy()
    top10["rank"] = top10.groupby("date")["ranking_score"].rank(method="first", ascending=False).astype(int)
    top10 = top10[top10["rank"] <= 10].sort_values(["date", "rank"]).reset_index(drop=True)
    return top10.loc[:, TOP10_COLUMNS].copy()


def add_evaluation_columns(df: pd.DataFrame) -> None:
    """Add actual performance evaluation display columns in-place."""
    actual_return = ((1 + df["target_gap"]) * (1 + df["target_intraday"]) - 1) * 100
    expected_return = df["expected_return"] * 100
    prediction_error = expected_return - actual_return
    hit = (
        ((expected_return > 0) & (actual_return > 0))
        | ((expected_return < 0) & (actual_return < 0))
    )
    abs_error = prediction_error.abs()

    df["Actual Return(%)"] = actual_return.round(2).map(lambda value: f"{value:.2f}%")
    df["Prediction Error(%)"] = prediction_error.round(2).map(lambda value: f"{value:.2f}%")
    df["Hit"] = hit.map({True: "\u2713", False: "\u2717"})
    df["Hit Binary"] = hit.astype(int)
    df["Prediction Quality"] = pd.cut(
        abs_error,
        bins=[-float("inf"), 1, 3, 5, float("inf")],
        labels=["Excellent", "Good", "Fair", "Poor"],
    ).astype(str)


def save_xlsx(path: Path, df: pd.DataFrame, sheet_name: str) -> None:
    """Save a readable Excel workbook with filters, formatting, and explanation."""
    with pd.ExcelWriter(path) as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        worksheet = writer.sheets[sheet_name]
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        style_header(worksheet)
        format_price_columns(worksheet)
        add_conditional_formatting(worksheet, df)
        add_evaluation_conditional_formatting(worksheet, df)
        autofit_columns(worksheet)
        add_explanation_sheet(writer.book)


def style_header(worksheet: object) -> None:
    """Apply simple header styling."""
    fill = PatternFill("solid", fgColor="1F2937")
    font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = fill
        cell.font = font


def add_conditional_formatting(worksheet: object, df: pd.DataFrame) -> None:
    """Add AI Score and Expected Return conditional formatting."""
    if df.empty:
        return
    headers = {cell.value: cell.column for cell in worksheet[1]}
    max_row = len(df) + 1
    if "AI Score" in headers:
        column = get_column_letter(headers["AI Score"])
        cell_range = f"{column}2:{column}{max_row}"
        worksheet.conditional_formatting.add(
            cell_range,
            CellIsRule(operator="between", formula=["90", "100"], fill=PatternFill("solid", fgColor="006100")),
        )
        worksheet.conditional_formatting.add(
            cell_range,
            CellIsRule(operator="between", formula=["80", "89.999"], fill=PatternFill("solid", fgColor="C6EFCE")),
        )
        worksheet.conditional_formatting.add(
            cell_range,
            CellIsRule(operator="between", formula=["70", "79.999"], fill=PatternFill("solid", fgColor="FFEB9C")),
        )
        worksheet.conditional_formatting.add(
            cell_range,
            CellIsRule(operator="between", formula=["60", "69.999"], fill=PatternFill("solid", fgColor="F4B183")),
        )
        worksheet.conditional_formatting.add(
            cell_range,
            CellIsRule(operator="lessThan", formula=["60"], fill=PatternFill("solid", fgColor="FCE4D6")),
        )
    if "Expected Return(%)" in headers:
        column = get_column_letter(headers["Expected Return(%)"])
        cell_range = f"{column}2:{column}{max_row}"
        worksheet.conditional_formatting.add(
            cell_range,
            FormulaRule(formula=[f'VALUE(SUBSTITUTE(${column}2,"%",""))>0'], font=Font(color="0000FF")),
        )
        worksheet.conditional_formatting.add(
            cell_range,
            FormulaRule(formula=[f'VALUE(SUBSTITUTE(${column}2,"%",""))<0'], font=Font(color="FF0000")),
        )


def format_price_columns(worksheet: object) -> None:
    """Apply integer KRW display format to price columns."""
    headers = {cell.value: cell.column for cell in worksheet[1]}
    for header in PRICE_COLUMNS:
        if header not in headers:
            continue
        column_letter = get_column_letter(headers[header])
        for cell in worksheet[f"{column_letter}"][1:]:
            cell.number_format = "#,##0"


def add_evaluation_conditional_formatting(worksheet: object, df: pd.DataFrame) -> None:
    """Add conditional formatting for hit, prediction error, and quality."""
    if df.empty:
        return
    headers = {cell.value: cell.column for cell in worksheet[1]}
    max_row = len(df) + 1
    if "Hit" in headers:
        column = get_column_letter(headers["Hit"])
        cell_range = f"{column}2:{column}{max_row}"
        worksheet.conditional_formatting.add(
            cell_range,
            FormulaRule(formula=[f'${column}2="\u2713"'], fill=PatternFill("solid", fgColor="C6EFCE")),
        )
        worksheet.conditional_formatting.add(
            cell_range,
            FormulaRule(formula=[f'${column}2="\u2717"'], fill=PatternFill("solid", fgColor="FCE4D6")),
        )
    if "Prediction Error(%)" in headers:
        column = get_column_letter(headers["Prediction Error(%)"])
        cell_range = f"{column}2:{column}{max_row}"
        abs_formula = f'ABS(VALUE(SUBSTITUTE(${column}2,"%","")))'
        worksheet.conditional_formatting.add(
            cell_range,
            FormulaRule(formula=[f"{abs_formula}<=1"], fill=PatternFill("solid", fgColor="006100")),
        )
        worksheet.conditional_formatting.add(
            cell_range,
            FormulaRule(formula=[f"AND({abs_formula}>1,{abs_formula}<=3)"], fill=PatternFill("solid", fgColor="C6EFCE")),
        )
        worksheet.conditional_formatting.add(
            cell_range,
            FormulaRule(formula=[f"AND({abs_formula}>3,{abs_formula}<=5)"], fill=PatternFill("solid", fgColor="FFEB9C")),
        )
        worksheet.conditional_formatting.add(
            cell_range,
            FormulaRule(formula=[f"{abs_formula}>5"], fill=PatternFill("solid", fgColor="F4B183")),
        )
    if "Prediction Quality" in headers:
        column = get_column_letter(headers["Prediction Quality"])
        cell_range = f"{column}2:{column}{max_row}"
        colors = {
            "Excellent": "006100",
            "Good": "C6EFCE",
            "Fair": "FFEB9C",
            "Poor": "FCE4D6",
        }
        for label, color in colors.items():
            worksheet.conditional_formatting.add(
                cell_range,
                FormulaRule(formula=[f'${column}2="{label}"'], fill=PatternFill("solid", fgColor=color)),
            )


def autofit_columns(worksheet: object) -> None:
    """Set practical column widths."""
    for column_cells in worksheet.columns:
        header = str(column_cells[0].value)
        sample_lengths = [len(str(cell.value)) for cell in column_cells[:200] if cell.value is not None]
        width = min(max([len(header), *sample_lengths]) + 2, 32)
        worksheet.column_dimensions[column_cells[0].column_letter].width = width


def add_explanation_sheet(workbook: object) -> None:
    """Add an explanation sheet for AI display fields."""
    if "Explanation" in workbook.sheetnames:
        del workbook["Explanation"]
    sheet = workbook.create_sheet("Explanation")
    rows = [
        ["Field", "Meaning"],
        ["AI Score", "0-100 normalized output of Ranking LightGBM"],
        ["AI Score", "Higher = stronger model preference"],
        ["AI Score", "Not expected return"],
        ["AI Percentile", "Top x% within the same prediction date"],
        ["Top 1%", "Highest-ranked stocks"],
        ["Top 5%", "Very strong candidates"],
        ["Top 10%", "Strong candidates"],
        ["Top 20%", "Candidate watchlist"],
        ["Above Top 50%", "Generally not selected"],
        ["AI Rank", "Position among all stocks on the same prediction date"],
    ]
    for row in rows:
        sheet.append(row)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    style_header(sheet)
    sheet.column_dimensions["A"].width = 22
    sheet.column_dimensions["B"].width = 64


def percent_display_mean(series: pd.Series) -> float:
    """Convert percentage display strings back to numeric percentage points."""
    numeric = series.astype(str).str.rstrip("%").replace("", pd.NA).astype(float)
    return float(numeric.mean())


def percent_display_values(series: pd.Series) -> pd.Series:
    """Convert percentage display strings to percentage-point floats."""
    return series.astype(str).str.rstrip("%").replace("", pd.NA).astype(float)


def percentile_distribution(readable: pd.DataFrame) -> dict[str, int]:
    """Count rows at common Top x% thresholds."""
    percentile = readable["AI Percentile"].str.extract(r"Top (\d+)%")[0].astype(int)
    return {
        "Top 1%": int((percentile <= 1).sum()),
        "Top 5%": int((percentile <= 5).sum()),
        "Top 10%": int((percentile <= 10).sum()),
        "Top 20%": int((percentile <= 20).sum()),
        "Top 50%": int((percentile <= 50).sum()),
    }


def render_summary(readable: pd.DataFrame, top10: pd.DataFrame) -> str:
    """Render Markdown summary for readable reports."""
    unknown_count = int(readable["ticker_name"].eq("UNKNOWN").sum())
    samsung_rows = readable[readable["ticker"].eq("005930")]
    samsung_mapped = bool(not samsung_rows.empty and samsung_rows["ticker_name"].iloc[0] == SAMSUNG_NAME)
    top20 = (
        top10.groupby(["ticker", "ticker_name"])
        .size()
        .reset_index(name="top10_appearances")
        .sort_values(["top10_appearances", "ticker"], ascending=[False, True])
        .head(20)
    )
    actual_return = percent_display_values(readable["Actual Return(%)"])
    prediction_error = percent_display_values(readable["Prediction Error(%)"])
    top10_actual_return = percent_display_values(top10["Actual Return(%)"])
    top10_prediction_error = percent_display_values(top10["Prediction Error(%)"])
    hit_ratio = float(readable["Hit Binary"].mean() * 100)
    top10_hit_ratio = float(top10["Hit Binary"].mean() * 100)
    evaluation_frame = readable.assign(
        _actual_return=actual_return,
        _abs_error=prediction_error.abs(),
    )
    daily_performance = evaluation_frame.groupby("date").agg(
        average_actual_return=("_actual_return", "mean"),
        mean_absolute_error=("_abs_error", "mean"),
    )
    best_day = daily_performance.sort_values(
        ["average_actual_return", "mean_absolute_error"],
        ascending=[False, True],
    ).index[0]
    worst_day = daily_performance.sort_values(
        ["average_actual_return", "mean_absolute_error"],
        ascending=[True, False],
    ).index[0]
    stock_accuracy = (
        evaluation_frame.groupby(["ticker", "ticker_name"])
        .agg(mean_absolute_error=("_abs_error", "mean"), rows=("ticker", "size"))
        .reset_index()
    )
    most_accurate = stock_accuracy.sort_values(["mean_absolute_error", "ticker"]).head(20)
    least_accurate = stock_accuracy.sort_values(
        ["mean_absolute_error", "ticker"],
        ascending=[False, True],
    ).head(20)
    distribution = percentile_distribution(readable)
    return "\n".join(
        [
            "# Full Universe Validation Predictions Summary",
            "",
            f"- Prediction rows: `{len(readable)}`",
            f"- Date range: `{readable['date'].min().date()}` to `{readable['date'].max().date()}`",
            f"- Unique dates: `{readable['date'].nunique()}`",
            f"- Unique tickers: `{readable['ticker'].nunique()}`",
            f"- Unknown ticker count: `{unknown_count}`",
            f"- 005930 exists: `{bool(not samsung_rows.empty)}`",
            f"- 005930 maps to Samsung Electronics: `{samsung_mapped}`",
            f"- Total Top10 rows: `{len(top10)}`",
            f"- AI Score min: `{readable['AI Score'].min():.1f}`",
            f"- AI Score max: `{readable['AI Score'].max():.1f}`",
            f"- AI Score average: `{readable['AI Score'].mean():.1f}`",
            f"- Top10 average AI Score: `{top10['AI Score'].mean():.1f}`",
            f"- Top10 average Expected Return(%): `{percent_display_mean(top10['Expected Return(%)']):.2f}%`",
            f"- Top10 average Gap Forecast(%): `{percent_display_mean(top10['Gap Forecast(%)']):.2f}%`",
            f"- Top10 average Intraday Forecast(%): `{percent_display_mean(top10['Intraday Forecast(%)']):.2f}%`",
            f"- Average Actual Return(%): `{actual_return.mean():.2f}%`",
            f"- Average Prediction Error(%): `{prediction_error.mean():.2f}%`",
            f"- Mean Absolute Error(%): `{prediction_error.abs().mean():.2f}%`",
            f"- Hit Ratio(%): `{hit_ratio:.2f}%`",
            f"- Top10 Hit Ratio(%): `{top10_hit_ratio:.2f}%`",
            f"- Top10 Average Actual Return(%): `{top10_actual_return.mean():.2f}%`",
            f"- Top10 Average Prediction Error(%): `{top10_prediction_error.mean():.2f}%`",
            f"- Best prediction day: `{pd.Timestamp(best_day).date()}`",
            f"- Worst prediction day: `{pd.Timestamp(worst_day).date()}`",
            "",
            "## Distribution",
            "",
            f"- Top 1%: `{distribution['Top 1%']}`",
            f"- Top 5%: `{distribution['Top 5%']}`",
            f"- Top 10%: `{distribution['Top 10%']}`",
            f"- Top 20%: `{distribution['Top 20%']}`",
            f"- Top 50%: `{distribution['Top 50%']}`",
            "",
            "## Top 20 Most Frequently Selected Tickers In Daily Top10",
            "",
            top20.to_markdown(index=False),
            "",
            "## Top20 Most Accurate Stocks",
            "",
            most_accurate.to_markdown(index=False),
            "",
            "## Top20 Least Accurate Stocks",
            "",
            least_accurate.to_markdown(index=False),
            "",
        ]
    )


def main() -> None:
    """Generate readable full-universe validation reports."""
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    print("Loading predictions and ticker names...")
    predictions = load_predictions()
    ticker_names = load_ticker_names()
    print("Creating readable and Top10 reports...")
    readable = create_readable(predictions, ticker_names)
    top10 = create_top10(readable)

    readable.to_csv(READABLE_CSV, index=False, encoding="utf-8-sig")
    top10.to_csv(TOP10_CSV, index=False, encoding="utf-8-sig")
    save_xlsx(READABLE_XLSX, readable, "Predictions")
    save_xlsx(TOP10_XLSX, top10, "Top10ByDate")
    SUMMARY_MD.write_text(render_summary(readable, top10), encoding="utf-8")

    print("Readable full-universe report complete")
    print(f"Total rows: {len(readable)}")
    print(f"Top10 rows: {len(top10)}")
    print(f"Unknown ticker count: {int(readable['ticker_name'].eq('UNKNOWN').sum())}")
    print("Formatted columns: ['AI Score', 'AI Percentile', 'AI Rank', 'Expected Return(%)', 'Gap Forecast(%)', 'Intraday Forecast(%)', 'Actual Gap(%)', 'Actual Intraday(%)']")
    print(f"Price columns rounded to integer KRW: {PRICE_COLUMNS}")


if __name__ == "__main__":
    main()
