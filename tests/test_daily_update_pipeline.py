"""Tests for Daily Update Pipeline Part 1A."""

from __future__ import annotations

import json
import os
import subprocess
import sys
import threading
from types import SimpleNamespace
from pathlib import Path

import pandas as pd
import pytest

import scripts.run_daily_update_pipeline as daily_pipeline
import scripts.diagnose_pykrx_daily_download as diagnose_download
from src.pipeline.archive import copy_status_into_archive, create_daily_archive, verify_archive_integrity
from src.pipeline.config import DIRECTORY_KEYS, load_daily_update_config
from src.pipeline.daily_model import build_daily_model_spec, train_daily_models
from src.pipeline.daily_prediction import build_daily_prediction_frame
from src.pipeline.daily_report import enrich_predictions_for_report, generate_daily_top10_report
from src.pipeline.daily_context import DailyRunContext, build_daily_run_context, fallback_to_latest_clean_context
from src.pipeline.env import load_project_env
from src.pipeline.feature_update import (
    optimize_feature_frame,
    run_feature_update,
    safe_append_features,
    validate_required_feature_availability,
)
from src.pipeline.feature_source_completeness import FeatureSourceCompletenessChecker
from src.pipeline.macro_update import (
    build_macro_update_row,
    run_macro_update,
    safe_append_macro,
)
from src.pipeline.macro_download import (
    MACRO_TICKERS,
    MacroDataUnavailableError,
    append_latest_macro,
    download_required_macro_row,
    print_macro_download_failure,
    print_macro_download_success,
    run_production_macro_download,
    validate_macro_source_frame,
)
from src.pipeline.market_data_update import (
    call_with_timeout,
    download_daily_ohlcv,
    download_with_retries,
    normalize_pykrx_ohlcv,
    run_ohlcv_update,
    safe_append_ohlcv,
    split_valid_invalid_ohlcv,
)
from src.pipeline.rolling_window import get_model_feature_columns, select_rolling_train_window
from src.pipeline.status import DailyUpdateStatus, write_status
from src.pipeline.training_update import (
    TrainingUpdateResult,
    build_target_available_rows,
    count_leakage_violations,
    run_training_update,
    safe_append_training_rows,
)


PROJECT_ROOT = Path(__file__).resolve().parents[1]
CONFIG_PATH = PROJECT_ROOT / "configs" / "daily_update.yaml"
SCRIPT_PATH = PROJECT_ROOT / "scripts" / "run_daily_update_pipeline.py"


def pykrx_test_python() -> str:
    """Return the interpreter that should be used for pykrx import probes."""
    venv_python = PROJECT_ROOT / ".venv" / "Scripts" / "python.exe"
    return str(venv_python if venv_python.exists() else Path(sys.executable))


class _PassingFeatureSourceCompletenessChecker:
    """Default passing checker for pipeline integration tests."""

    def __init__(self, _config, expected_date):
        self.expected_date = pd.Timestamp(expected_date).date().isoformat()

    def check(self) -> dict[str, object]:
        result: dict[str, object] = {
            "expected_date": self.expected_date,
            "all_available": True,
            "failure_reason": "",
            "failed_sources": [],
        }
        for source in ["krx", "nasdaq", "sox", "sp500", "vix", "wti", "usdkrw", "us10y", "gold", "dxy"]:
            result[f"{source}_check_enabled"] = source not in {"us10y", "gold", "dxy"}
            result[f"{source}_available"] = True
            result[f"actual_{source}_date"] = self.expected_date
        result["us10y_skipped_reason"] = "disabled in config"
        result["gold_skipped_reason"] = "disabled in config"
        result["dxy_skipped_reason"] = "disabled in config"
        return result


@pytest.fixture(autouse=True)
def default_pipeline_source_check_passes(monkeypatch):
    """Keep existing mocked pipeline tests focused unless they override source checks."""
    monkeypatch.setattr(
        daily_pipeline,
        "FeatureSourceCompletenessChecker",
        _PassingFeatureSourceCompletenessChecker,
    )


def _mock_ohlcv_result() -> SimpleNamespace:
    return SimpleNamespace(
        universe_count=2,
        raw_rows_downloaded_or_found=2,
        raw_rows_added=0,
        cleaned_rows_added=0,
        invalid_ohlcv_rows=0,
        invalid_ohlcv_tickers=[],
        missing_005930=False,
        daily_ohlcv_snapshot_paths=[],
        ohlcv_download_mode="no_download",
        ohlcv_download_attempts=0,
        ohlcv_download_timed_out=False,
        ohlcv_download_failed=False,
        ohlcv_download_error=None,
        pykrx_rows_returned=2,
        pykrx_empty_response=False,
        pykrx_missing_columns=False,
        pykrx_data_unavailable=False,
        pykrx_download_method="per_ticker",
        pykrx_tickers_requested=2,
        pykrx_tickers_downloaded=2,
        pykrx_tickers_failed=0,
        pykrx_failed_tickers_sample=[],
        pykrx_success_ratio=1.0,
        trading_value_estimated_count=0,
        used_existing_data_fallback=False,
        warnings=[],
        errors=[],
    )


def _mock_fallback_ohlcv_result() -> SimpleNamespace:
    result = _mock_ohlcv_result()
    return SimpleNamespace(
        **{
            **result.__dict__,
            "raw_rows_downloaded_or_found": 0,
            "ohlcv_download_mode": "no_download",
            "pykrx_rows_returned": 0,
            "pykrx_empty_response": True,
            "pykrx_missing_columns": False,
            "pykrx_data_unavailable": True,
            "pykrx_tickers_downloaded": 0,
            "pykrx_tickers_failed": 2,
            "pykrx_failed_tickers_sample": ["005930", "000660"],
            "pykrx_success_ratio": 0.0,
            "used_existing_data_fallback": True,
            "warnings": ["skip_download_enabled_no_ohlcv_rows_found_for_2026-06-15"],
        }
    )


def _mock_macro_result() -> SimpleNamespace:
    return SimpleNamespace(
        macro_update_mode="existing",
        macro_source_date="2026-06-12",
        macro_rows_added=0,
        macro_missing_after_update={},
        daily_macro_snapshot_path=None,
        warnings=[],
        errors=[],
    )


def _mock_feature_result() -> SimpleNamespace:
    return SimpleNamespace(
        feature_rows_added=0,
        feature_rows_replaced=0,
        daily_feature_snapshot_path=None,
        feature_update_mode="existing",
        feature_update_date="2026-06-12",
        feature_missing_count=0,
        feature_column_count=1,
        feature_ticker_count=2,
        warnings=[],
        errors=[],
    )


def _mock_training_result() -> TrainingUpdateResult:
    return TrainingUpdateResult(
        training_rows_added=0,
        training_rows_replaced=0,
        daily_training_snapshot_path=None,
        training_update_mode="existing",
        target_feature_dates_added=[],
        leakage_violations=0,
        forbidden_model_features_found=[],
        rolling_train_start_date="2025-01-01",
        rolling_train_end_date="2025-12-31",
        rolling_train_unique_dates=250,
        rolling_train_rows=500,
        selected_feature_count=1,
    )


def _mock_model_result() -> SimpleNamespace:
    return SimpleNamespace(
        model_bundle=None,
        train_df=pd.DataFrame(),
        feature_columns=["feature_a"],
        train_start_date="2025-01-01",
        train_end_date="2025-12-31",
        rolling_train_days=250,
        rolling_train_rows=500,
        model_output_dir=str(PROJECT_ROOT / "outputs" / "daily_models" / "20990201"),
        model_paths={},
    )


def _mock_prediction_result() -> SimpleNamespace:
    return SimpleNamespace(
        prediction_df=pd.DataFrame(),
        prediction_output_csv=str(PROJECT_ROOT / "outputs" / "daily_predictions" / "predictions_20990202.csv"),
        prediction_output_parquet=str(PROJECT_ROOT / "outputs" / "daily_predictions" / "predictions_20990202.parquet"),
        prediction_rows=2,
    )


def _mock_report_result() -> SimpleNamespace:
    return SimpleNamespace(
        top10_report_csv=str(PROJECT_ROOT / "reports" / "daily" / "top10_20990202.csv"),
        top10_report_xlsx=str(PROJECT_ROOT / "reports" / "daily" / "top10_20990202.xlsx"),
        daily_summary_report=str(PROJECT_ROOT / "reports" / "daily" / "daily_update_summary_20990202.md"),
        top10_tickers=["005930"],
        top10_ticker_names=["Samsung"],
        top10_average_ai_score=95.0,
        top10_average_expected_return=0.01,
        warnings=[],
    )


def test_daily_update_config_file_exists_and_loads() -> None:
    """Config file should exist and expose required values."""
    assert CONFIG_PATH.exists()
    config = load_daily_update_config(CONFIG_PATH)
    assert config.rolling_train_days == 350
    assert config.values["recommended_run_time"] == "08:30"
    assert config.values["market_open_time"] == "09:00"


def test_missing_env_does_not_crash(tmp_path: Path, monkeypatch) -> None:
    """Missing .env should not matter for pykrx-only daily updates."""
    monkeypatch.delenv("KRX_ID", raising=False)
    monkeypatch.delenv("KRX_PW", raising=False)

    warnings = load_project_env(tmp_path)

    assert warnings == []


def test_fake_env_ignores_legacy_credentials_without_printing_values(tmp_path: Path, monkeypatch) -> None:
    """Legacy KRX credential keys should be ignored without exposing values."""
    monkeypatch.delenv("KRX_ID", raising=False)
    monkeypatch.delenv("KRX_PW", raising=False)
    secret_id = "fake_user_secret"
    secret_pw = "fake_password_secret"
    (tmp_path / ".env").write_text(f"KRX_ID={secret_id}\nKRX_PW={secret_pw}\n", encoding="utf-8")

    warnings = load_project_env(tmp_path)

    assert warnings == []
    assert "KRX_ID" not in os.environ
    assert "KRX_PW" not in os.environ
    assert secret_id not in json.dumps(warnings)
    assert secret_pw not in json.dumps(warnings)


def test_missing_credentials_do_not_matter(monkeypatch) -> None:
    """KRX credentials are not required by the pykrx-only path."""
    monkeypatch.delenv("KRX_ID", raising=False)
    monkeypatch.delenv("KRX_PW", raising=False)

    warnings = load_project_env(PROJECT_ROOT / "path-that-does-not-exist")

    assert warnings == []


def test_credentials_not_written_to_status_or_logs(tmp_path: Path, monkeypatch) -> None:
    """Credential values must not appear in status JSON or logs."""
    secret_id = "secret_krx_id_value"
    secret_pw = "secret_krx_pw_value"
    (tmp_path / ".env").write_text(f"KRX_ID={secret_id}\nKRX_PW={secret_pw}\n", encoding="utf-8")
    config = load_daily_update_config(CONFIG_PATH)
    clean_path = tmp_path / "clean.parquet"
    valid_ohlcv_rows("2026-06-12").to_parquet(clean_path, index=False)
    values = dict(config.values)
    values["daily_status_dir"] = str(tmp_path / "status")
    values["log_dir"] = str(tmp_path / "logs")
    values["clean_ohlcv_file"] = str(clean_path)
    test_config = type(config)(config_path=config.config_path, values=values, project_root=tmp_path)
    monkeypatch.setattr(daily_pipeline, "load_daily_update_config", lambda _path: test_config)
    monkeypatch.setattr(daily_pipeline, "run_ohlcv_update", lambda **_kwargs: _mock_ohlcv_result())
    monkeypatch.setattr(daily_pipeline, "run_macro_update", lambda **_kwargs: _mock_macro_result())
    monkeypatch.setattr(daily_pipeline, "run_feature_update", lambda **_kwargs: _mock_feature_result())
    monkeypatch.setattr(daily_pipeline, "run_training_update", lambda **_kwargs: _mock_training_result())
    monkeypatch.setattr(daily_pipeline, "train_daily_models", lambda **_kwargs: _mock_model_result())
    monkeypatch.setattr(daily_pipeline, "generate_daily_predictions", lambda **_kwargs: _mock_prediction_result())
    monkeypatch.setattr(daily_pipeline, "generate_daily_top10_report", lambda **_kwargs: _mock_report_result())

    daily_pipeline.run_pipeline(
        daily_pipeline.parse_args(["--as-of-date", "2099-03-02", "--no-download", "--config", str(CONFIG_PATH)])
    )

    status_text = (tmp_path / "status" / "daily_update_status_20990302.json").read_text(encoding="utf-8")
    log_text = (tmp_path / "logs" / "daily_update_20990302.log").read_text(encoding="utf-8")
    assert secret_id not in status_text
    assert secret_pw not in status_text
    assert secret_id not in log_text
    assert secret_pw not in log_text


def test_cli_dry_run_runs_successfully() -> None:
    """Dry-run CLI should complete without writing status."""
    as_of_date = "2099-01-31"
    status_file = PROJECT_ROOT / "outputs" / "daily_status" / "daily_update_status_20990131.json"
    if status_file.exists():
        status_file.unlink()

    result = subprocess.run(
        [
            sys.executable,
            str(SCRIPT_PATH),
            "--as-of-date",
            as_of_date,
            "--dry-run",
            "--skip-download",
        ],
        cwd=PROJECT_ROOT,
        capture_output=True,
        text=True,
        check=False,
    )

    assert result.returncode == 0
    assert "Status JSON preview" in result.stdout
    assert not status_file.exists()


def test_normal_mode_creates_status_json(monkeypatch, tmp_path: Path) -> None:
    """Normal CLI should write the status JSON file."""
    as_of_date = "2099-02-01"
    config = load_daily_update_config(CONFIG_PATH)
    clean_path = tmp_path / "clean.parquet"
    valid_ohlcv_rows("2026-06-12").to_parquet(clean_path, index=False)
    values = dict(config.values)
    values["daily_status_dir"] = str(tmp_path / "status")
    values["log_dir"] = str(tmp_path / "logs")
    values["clean_ohlcv_file"] = str(clean_path)
    test_config = type(config)(config_path=config.config_path, values=values, project_root=PROJECT_ROOT)
    status_file = tmp_path / "status" / "daily_update_status_20990201.json"
    monkeypatch.setattr(daily_pipeline, "load_daily_update_config", lambda _path: test_config)
    monkeypatch.setattr(daily_pipeline, "run_ohlcv_update", lambda **_kwargs: _mock_ohlcv_result())
    monkeypatch.setattr(daily_pipeline, "run_macro_update", lambda **_kwargs: _mock_macro_result())
    monkeypatch.setattr(daily_pipeline, "run_feature_update", lambda **_kwargs: _mock_feature_result())
    monkeypatch.setattr(daily_pipeline, "run_training_update", lambda **_kwargs: _mock_training_result())
    monkeypatch.setattr(daily_pipeline, "train_daily_models", lambda **_kwargs: _mock_model_result())
    monkeypatch.setattr(daily_pipeline, "generate_daily_predictions", lambda **_kwargs: _mock_prediction_result())
    monkeypatch.setattr(daily_pipeline, "generate_daily_top10_report", lambda **_kwargs: _mock_report_result())

    daily_pipeline.run_pipeline(
        daily_pipeline.parse_args(["--as-of-date", as_of_date, "--skip-download", "--config", str(CONFIG_PATH)])
    )

    assert status_file.exists()


def test_status_json_contains_required_fields(monkeypatch, tmp_path: Path) -> None:
    """Written status JSON should contain all Part 1A required fields."""
    as_of_date = "2099-02-02"
    config = load_daily_update_config(CONFIG_PATH)
    values = dict(config.values)
    values["daily_status_dir"] = str(tmp_path / "status")
    values["log_dir"] = str(tmp_path / "logs")
    test_config = type(config)(config_path=config.config_path, values=values, project_root=PROJECT_ROOT)
    status_file = tmp_path / "status" / "daily_update_status_20990202.json"
    monkeypatch.setattr(daily_pipeline, "load_daily_update_config", lambda _path: test_config)
    monkeypatch.setattr(daily_pipeline, "run_ohlcv_update", lambda **_kwargs: _mock_ohlcv_result())
    monkeypatch.setattr(daily_pipeline, "run_macro_update", lambda **_kwargs: _mock_macro_result())
    monkeypatch.setattr(daily_pipeline, "run_feature_update", lambda **_kwargs: _mock_feature_result())
    monkeypatch.setattr(daily_pipeline, "run_training_update", lambda **_kwargs: _mock_training_result())
    monkeypatch.setattr(daily_pipeline, "train_daily_models", lambda **_kwargs: _mock_model_result())
    monkeypatch.setattr(daily_pipeline, "generate_daily_predictions", lambda **_kwargs: _mock_prediction_result())
    monkeypatch.setattr(daily_pipeline, "generate_daily_top10_report", lambda **_kwargs: _mock_report_result())

    daily_pipeline.run_pipeline(
        daily_pipeline.parse_args(["--as-of-date", as_of_date, "--force", "--skip-download", "--config", str(CONFIG_PATH)])
    )

    status = json.loads(status_file.read_text(encoding="utf-8"))
    required = {
        "run_timestamp",
        "run_date",
        "as_of_date",
        "latest_clean_data_date",
        "target_update_date",
        "update_date",
        "prediction_date",
        "dry_run",
        "skip_download",
        "force",
        "rolling_train_days",
        "production_mode",
        "pipeline_stop_reason",
        "pipeline_exit_code",
        "pipeline_exit_message",
        "latest_available_market_date",
        "stale_data_blocked_by_production_policy",
        "ohlcv_download_mode",
        "ohlcv_download_attempts",
        "ohlcv_download_timed_out",
        "ohlcv_download_failed",
        "ohlcv_download_error",
        "pykrx_rows_returned",
        "pykrx_empty_response",
        "pykrx_missing_columns",
        "pykrx_data_unavailable",
        "pykrx_download_method",
        "pykrx_tickers_requested",
        "pykrx_tickers_downloaded",
        "pykrx_tickers_failed",
        "pykrx_failed_tickers_sample",
        "pykrx_success_ratio",
        "trading_value_estimated_count",
        "used_existing_data_fallback",
        "attempted_download_date",
        "downloaded_update_date",
        "old_data_warning",
        "krx_login_attempted",
        "krx_login_success",
        "krx_login_timed_out",
        "krx_login_failed",
        "krx_login_error",
        "feature_source_completeness_passed",
        "expected_feature_date",
        "actual_krx_date",
        "actual_nasdaq_date",
        "actual_sox_date",
        "sox_check_enabled",
        "sox_close_present",
        "sox_return_present",
        "sox_return_non_null_count",
        "sox_failure_reason",
        "actual_sp500_date",
        "actual_vix_date",
        "actual_wti_date",
        "actual_usdkrw_date",
        "sources_using_prior_trading_day",
        "us_market_holiday_detected",
        "macro_date_policy",
        "macro_invalid_target_date_rows",
        "macro_source_actual_dates",
        "macro_source_expected_dates",
        "actual_us10y_date",
        "us10y_check_enabled",
        "actual_gold_date",
        "gold_check_enabled",
        "actual_dxy_date",
        "dxy_check_enabled",
        "failed_feature_sources",
        "feature_update_executed",
        "training_update_executed",
        "rolling_train_executed",
        "prediction_executed",
        "top10_generated",
        "top10_report_csv",
        "top10_report_xlsx",
        "daily_summary_report",
        "completed_steps",
        "warnings",
        "errors",
    }
    assert required <= set(status)
    assert status["as_of_date"] == as_of_date
    assert status["rolling_train_days"] == 350
    assert status["update_date"] is not None
    assert status["prediction_date"] is not None
    assert "status_written" in status["completed_steps"]
    assert status["top10_report_csv"].endswith(".csv")
    assert status["top10_report_xlsx"].endswith(".xlsx")
    assert status["daily_summary_report"].endswith(".md")
    assert "top10_report_generated" in status["completed_steps"]


def test_required_configured_directories_are_recognized() -> None:
    """Config should expose every required directory as a project path."""
    config = load_daily_update_config(CONFIG_PATH)
    directories = config.required_directories()

    assert set(DIRECTORY_KEYS) == set(directories)
    for path in directories.values():
        assert path.is_absolute()
        assert PROJECT_ROOT in (path, *path.parents)


def test_daily_run_context_creates_update_and_prediction_dates() -> None:
    """DailyRunContext should resolve update and prediction dates."""
    config = load_daily_update_config(CONFIG_PATH)
    context = build_daily_run_context(
        config=config,
        as_of_date="2026-06-16",
        dry_run=True,
        skip_download=True,
        force=False,
    )

    assert context.update_date <= context.as_of_date
    assert context.prediction_date > context.update_date
    assert context.latest_clean_data_date is not None


def test_daily_run_context_targets_previous_business_day_before_latest_clean_fallback() -> None:
    """Context should target previous business day before falling back to existing clean data."""
    config = load_daily_update_config(CONFIG_PATH)
    context = build_daily_run_context(
        config=config,
        as_of_date="2026-06-16",
        dry_run=True,
        skip_download=False,
        force=False,
    )

    assert context.target_update_date == "2026-06-15"
    assert context.update_date == "2026-06-15"
    assert context.prediction_date == "2026-06-16"


def test_failed_target_download_falls_back_to_latest_clean_date() -> None:
    """Failed target update should use latest clean data with OLD_DATA warning."""
    context = DailyRunContext(
        run_date="2026-06-16",
        as_of_date="2026-06-16",
        latest_clean_data_date="2026-06-12",
        target_update_date="2026-06-15",
        update_date="2026-06-15",
        prediction_date="2026-06-16",
        dry_run=False,
        skip_download=False,
        force=False,
        warnings=[],
    )

    fallback = fallback_to_latest_clean_context(context)

    assert fallback.update_date == "2026-06-12"
    assert fallback.prediction_date == "2026-06-15"
    assert any("OLD_DATA" in warning for warning in fallback.warnings)


def test_dry_run_status_preview_includes_update_and_prediction_dates() -> None:
    """Dry-run output should include resolved date fields."""
    result = subprocess.run(
        [sys.executable, str(SCRIPT_PATH), "--as-of-date", "2026-06-16", "--dry-run"],
        cwd=PROJECT_ROOT,
        capture_output=True,
        text=True,
        check=False,
    )

    assert result.returncode == 0
    assert '"update_date"' in result.stdout
    assert '"prediction_date"' in result.stdout


def sample_training_dataset(unique_dates: int = 260) -> pd.DataFrame:
    """Build a synthetic training dataset with multiple tickers per date."""
    dates = pd.bdate_range("2025-01-01", periods=unique_dates)
    rows = []
    for feature_date in dates:
        for ticker in ["005930", "000660"]:
            rows.append(
                {
                    "date": feature_date + pd.offsets.BDay(1),
                    "ticker": ticker,
                    "feature_date": feature_date,
                    "target_date": feature_date + pd.offsets.BDay(1),
                    "prediction_horizon": 1,
                    "prev_close": 100.0,
                    "feature_a": 1.0,
                    "target_ranking": 0.01,
                    "target_gap": 0.0,
                    "target_intraday": 0.01,
                }
            )
    return pd.DataFrame(rows)


def test_select_rolling_train_window_uses_exactly_250_unique_feature_dates() -> None:
    """Rolling window must select exactly 250 unique dates."""
    df = sample_training_dataset(260)
    prediction_date = pd.bdate_range("2025-01-01", periods=261)[-1]
    train_df, start, end, unique_dates, row_count = select_rolling_train_window(df, prediction_date, 250)

    assert len(unique_dates) == 250
    assert train_df["feature_date"].nunique() == 250
    assert row_count == len(train_df)
    assert start == pd.Timestamp(unique_dates[0])
    assert end == pd.Timestamp(unique_dates[-1])


def test_max_train_feature_date_is_before_prediction_date() -> None:
    """Selected train rows must be strictly before prediction date."""
    df = sample_training_dataset(260)
    prediction_date = pd.bdate_range("2025-01-01", periods=261)[-1]
    train_df, _, _, _, _ = select_rolling_train_window(df, prediction_date, 250)

    assert train_df["feature_date"].max() < prediction_date


def test_rolling_window_is_pure_rolling_not_expanding() -> None:
    """Old eligible dates should be dropped when more than 250 exist."""
    df = sample_training_dataset(260)
    prediction_date = pd.bdate_range("2025-01-01", periods=261)[-1]
    train_df, _, _, unique_dates, _ = select_rolling_train_window(df, prediction_date, 250)

    assert df["feature_date"].min() not in set(train_df["feature_date"])
    assert unique_dates[0] > df["feature_date"].min()


def test_rolling_window_raises_if_fewer_than_250_dates() -> None:
    """Insufficient date history should fail clearly."""
    df = sample_training_dataset(249)

    try:
        select_rolling_train_window(df, pd.Timestamp("2026-01-01"), 250)
    except ValueError as exc:
        assert "Need at least 250 unique feature_date" in str(exc)
    else:
        raise AssertionError("Expected ValueError for insufficient rolling dates")


def test_forbidden_columns_excluded_from_model_features() -> None:
    """Audit, identity, and target-prefixed columns should be excluded."""
    df = pd.DataFrame(
        {
            "date": [pd.Timestamp("2026-01-01")],
            "ticker": ["005930"],
            "ticker_name": ["Samsung"],
            "feature_date": [pd.Timestamp("2025-12-31")],
            "target_date": [pd.Timestamp("2026-01-01")],
            "prediction_horizon": [1],
            "prev_close": [100.0],
            "target_ranking": [0.01],
            "target_gap": [0.0],
            "target_intraday": [0.01],
            "target_extra": [0.02],
            "sector": ["Tech"],
            "market_type": ["KOSPI"],
            "market_cap_group": ["Large"],
            "feature_a": [1.0],
        }
    )

    features = get_model_feature_columns(df)

    assert features == ["feature_a"]
    assert "prev_close" not in features
    assert all(not column.startswith("target_") for column in features)


def valid_ohlcv_rows(date: str = "2026-06-12") -> pd.DataFrame:
    """Create valid synthetic OHLCV rows."""
    return pd.DataFrame(
        {
            "date": pd.to_datetime([date, date]),
            "ticker": ["005930", "000660"],
            "open": [100.0, 200.0],
            "high": [110.0, 220.0],
            "low": [90.0, 190.0],
            "close": [105.0, 210.0],
            "volume": [1000, 2000],
            "trading_value": [105000.0, 420000.0],
        }
    )


def make_source_check_config(tmp_path: Path, missing_source: str | None = None, macro_date: str = "2026-06-15"):
    """Build a temp config with synthetic KRX and macro source files."""
    config = load_daily_update_config(CONFIG_PATH)
    krx_path = tmp_path / "krx.parquet"
    macro_path = tmp_path / "macro.parquet"
    if missing_source != "krx":
        valid_ohlcv_rows("2026-06-15").to_parquet(krx_path, index=False)
    macro_columns = {
        "date": pd.to_datetime([macro_date]),
        "nasdaq_close": [100.0],
        "sox_close": [200.0],
        "sp500_close": [100.0],
        "vix_close": [20.0],
        "wti_close": [70.0],
        "usdkrw": [1300.0],
        "us10y": [4.0],
        "gold": [2300.0],
        "dxy": [105.0],
    }
    source_to_column = {
        "nasdaq": "nasdaq_close",
        "sox": "sox_close",
        "sp500": "sp500_close",
        "vix": "vix_close",
        "wti": "wti_close",
        "usdkrw": "usdkrw",
        "us10y": "us10y",
        "gold": "gold",
        "dxy": "dxy",
    }
    if missing_source in source_to_column:
        macro_columns.pop(source_to_column[missing_source])
    pd.DataFrame(macro_columns).to_parquet(macro_path, index=False)
    values = dict(config.values)
    values["clean_ohlcv_file"] = str(krx_path)
    values["macro_file"] = str(macro_path)
    values["feature_file"] = str(tmp_path / "features_not_built.parquet")
    values["enable_us10y_check"] = False
    values["enable_gold_check"] = False
    values["enable_dxy_check"] = False
    return type(config)(config_path=config.config_path, values=values, project_root=PROJECT_ROOT)


@pytest.mark.parametrize("missing_source", ["krx", "nasdaq", "sox", "sp500", "vix", "wti", "usdkrw"])
def test_feature_source_completeness_required_source_missing(tmp_path: Path, missing_source: str) -> None:
    """Missing required source should fail completeness."""
    config = make_source_check_config(tmp_path, missing_source=missing_source)

    result = FeatureSourceCompletenessChecker(config, "2026-06-15").check()

    assert result[f"{missing_source}_available"] is False
    assert result["all_available"] is False
    assert missing_source in result["failure_reason"]


def test_feature_source_completeness_us10y_missing_does_not_fail_by_default(tmp_path: Path) -> None:
    """US10Y should be skipped by default and should not block production completeness."""
    config = make_source_check_config(tmp_path, missing_source="us10y")

    result = FeatureSourceCompletenessChecker(config, "2026-06-15").check()

    assert result["us10y_check_enabled"] is False
    assert result["us10y_available"] is True
    assert result["us10y_skipped_reason"] == "disabled in config"
    assert result["all_available"] is True
    assert result["failed_sources"] == []


@pytest.mark.parametrize("missing_source", ["gold", "dxy"])
def test_feature_source_completeness_optional_sources_missing_do_not_fail_by_default(
    tmp_path: Path,
    missing_source: str,
) -> None:
    """Gold and DXY should be skipped by default and should not block production completeness."""
    config = make_source_check_config(tmp_path, missing_source=missing_source)

    result = FeatureSourceCompletenessChecker(config, "2026-06-15").check()

    assert result[f"{missing_source}_check_enabled"] is False
    assert result[f"{missing_source}_available"] is True
    assert result[f"{missing_source}_skipped_reason"] == "disabled in config"
    assert result["all_available"] is True
    assert result["failed_sources"] == []


def test_feature_source_completeness_date_mismatch_fails(tmp_path: Path) -> None:
    """A US source older than tolerance should fail even if columns exist."""
    config = make_source_check_config(tmp_path, macro_date="2026-06-09")

    result = FeatureSourceCompletenessChecker(config, "2026-06-15").check()

    assert result["nasdaq_available"] is False
    assert result["actual_nasdaq_date"] == "2026-06-09"
    assert result["all_available"] is False


def test_feature_source_completeness_all_pass(tmp_path: Path) -> None:
    """All configured sources on expected date should pass."""
    config = make_source_check_config(tmp_path)

    result = FeatureSourceCompletenessChecker(config, "2026-06-15").check()

    assert result["all_available"] is True
    assert result["failure_reason"] == ""
    assert result["actual_krx_date"] == "2026-06-15"
    assert result["actual_sox_date"] == "2026-06-15"
    assert result["sox_check_enabled"] is True


def test_sox_source_check_passes_matching_date_and_valid_return(tmp_path: Path) -> None:
    """SOX should pass when its close is current and its daily return is usable."""
    config = make_source_check_config(tmp_path)
    feature_path = tmp_path / "features.parquet"
    pd.DataFrame(
        {
            "date": pd.to_datetime(["2026-06-15", "2026-06-15"]),
            "ticker": ["005930", "000660"],
            "sox_return_1d": [0.01, 0.01],
        }
    ).to_parquet(feature_path, index=False)
    values = dict(config.values)
    values["feature_file"] = str(feature_path)
    test_config = type(config)(config_path=config.config_path, values=values, project_root=PROJECT_ROOT)

    result = FeatureSourceCompletenessChecker(test_config, "2026-06-15").check()

    assert result["sox_available"] is True
    assert result["all_available"] is True
    assert result["sox_close_present"] is True
    assert result["sox_return_present"] is True
    assert result["sox_return_non_null_count"] == 2
    assert result["sox_failure_reason"] is None


def test_sox_source_check_fails_matching_date_when_return_is_all_null(tmp_path: Path) -> None:
    """A matching SOX close date must not hide an unusable return feature."""
    config = make_source_check_config(tmp_path)
    feature_path = tmp_path / "features.parquet"
    pd.DataFrame(
        {
            "date": pd.to_datetime(["2026-06-15", "2026-06-15"]),
            "ticker": ["005930", "000660"],
            "sox_return_1d": [float("nan"), float("nan")],
        }
    ).to_parquet(feature_path, index=False)
    values = dict(config.values)
    values["feature_file"] = str(feature_path)
    test_config = type(config)(config_path=config.config_path, values=values, project_root=PROJECT_ROOT)

    result = FeatureSourceCompletenessChecker(test_config, "2026-06-15").check()

    expected_reason = "SOX close exists but sox_return_1d cannot be computed yet because prior SOX history is missing."
    assert result["actual_sox_date"] == result["expected_sox_date"] == "2026-06-15"
    assert result["sox_available"] is False
    assert result["sox_return_non_null_count"] == 0
    assert result["sox_failure_reason"] == expected_reason
    assert expected_reason in result["failure_reason"]
    assert result["failure_reason"] != "sox: expected 2026-06-15, actual 2026-06-15"


def test_sox_source_check_accepts_case_insensitive_columns(tmp_path: Path) -> None:
    """External SOX column casing should not create a false missing-source failure."""
    config = make_source_check_config(tmp_path)
    macro_path = tmp_path / "macro_upper.parquet"
    pd.DataFrame(
        {
            "date": pd.to_datetime(["2026-06-15"]),
            "nasdaq_close": [100.0],
            "SOX_CLOSE": [200.0],
            "ACTUAL_SOX_DATE": pd.to_datetime(["2026-06-15"]),
            "sp500_close": [100.0],
            "vix_close": [20.0],
            "wti_close": [70.0],
            "usdkrw": [1300.0],
        }
    ).to_parquet(macro_path, index=False)
    values = dict(config.values)
    values["macro_file"] = str(macro_path)
    test_config = type(config)(config_path=config.config_path, values=values, project_root=PROJECT_ROOT)

    result = FeatureSourceCompletenessChecker(test_config, "2026-06-15").check()

    assert result["sox_available"] is True
    assert result["actual_sox_date"] == "2026-06-15"


def test_progress_output_contains_step_total(capsys) -> None:
    """Progress output should show current step and total step count."""
    daily_pipeline.emit_step_progress(3, "KRX Download", "START", 0.0)

    output = capsys.readouterr().out
    assert "[03/12] KRX Download" in output
    assert "Progress:" in output
    assert "Status: START" in output


def test_heartbeat_message_can_be_emitted(capsys) -> None:
    """Heartbeat output should show current step and safe interrupt hint."""
    daily_pipeline.emit_heartbeat("Model Training", 90.2)

    output = capsys.readouterr().out
    assert "Still running..." in output
    assert "Current step: Model Training" in output
    assert "Press Ctrl+C to stop safely." in output


def test_keyboard_interrupt_returns_130(monkeypatch, capsys) -> None:
    """KeyboardInterrupt should return the conventional 130 exit code."""
    monkeypatch.setattr(daily_pipeline, "CURRENT_STATUS", None)
    monkeypatch.setattr(daily_pipeline, "CURRENT_CONFIG", None)
    monkeypatch.setattr(daily_pipeline, "run_pipeline", lambda _args: (_ for _ in ()).throw(KeyboardInterrupt()))

    exit_code = daily_pipeline.main(["--as-of-date", "2099-05-01"])

    assert exit_code == 130
    assert "Interrupted by user" in capsys.readouterr().out


def test_generic_exception_returns_1(monkeypatch, capsys) -> None:
    """Unexpected exceptions should return failure code 1 with visible reason."""
    monkeypatch.setattr(daily_pipeline, "CURRENT_STATUS", None)
    monkeypatch.setattr(daily_pipeline, "CURRENT_CONFIG", None)
    monkeypatch.setattr(daily_pipeline, "run_pipeline", lambda _args: (_ for _ in ()).throw(RuntimeError("boom")))
    monkeypatch.setattr(daily_pipeline.time, "sleep", lambda _seconds: pytest.fail("generic failure should not sleep"))

    exit_code = daily_pipeline.main(["--as-of-date", "2099-05-02"])

    output = capsys.readouterr().out
    assert exit_code == 1
    assert "Pipeline FAILED" in output
    assert "boom" in output


class _FailingFeatureSourceCompletenessChecker:
    """Configurable failing checker for production stop tests."""

    missing_source = "krx"

    def __init__(self, _config, expected_date):
        self.expected_date = pd.Timestamp(expected_date).date().isoformat()

    def check(self) -> dict[str, object]:
        sox_reason = "SOX close exists but sox_return_1d cannot be computed yet because prior SOX history is missing."
        failure_detail = sox_reason if self.missing_source == "sox" else f"expected {self.expected_date}, actual missing"
        result: dict[str, object] = {
            "expected_date": self.expected_date,
            "all_available": False,
            "failure_reason": f"{self.missing_source}: {failure_detail}",
            "failed_sources": [self.missing_source],
            "sox_close_present": True,
            "sox_return_present": True,
            "sox_return_non_null_count": 0 if self.missing_source == "sox" else 2,
            "sox_failure_reason": sox_reason if self.missing_source == "sox" else None,
        }
        for source in ["krx", "nasdaq", "sox", "sp500", "vix", "wti", "usdkrw", "us10y", "gold", "dxy"]:
            result[f"{source}_check_enabled"] = source not in {"us10y", "gold", "dxy"}
            result[f"{source}_available"] = source != self.missing_source
            result[f"actual_{source}_date"] = (
                self.expected_date if source == "sox" else None
            ) if source == self.missing_source else self.expected_date
        result["us10y_available"] = True
        result["us10y_skipped_reason"] = "disabled in config"
        result["gold_available"] = True
        result["gold_skipped_reason"] = "disabled in config"
        result["dxy_available"] = True
        result["dxy_skipped_reason"] = "disabled in config"
        return result


def _pipeline_test_config(tmp_path: Path):
    config = load_daily_update_config(CONFIG_PATH)
    values = dict(config.values)
    values["daily_status_dir"] = str(tmp_path / "status")
    values["log_dir"] = str(tmp_path / "logs")
    values["daily_report_dir"] = str(tmp_path / "reports")
    return type(config)(config_path=config.config_path, values=values, project_root=PROJECT_ROOT)


@pytest.mark.parametrize(
    ("missing_source", "skipped_step"),
    [
        ("krx", "feature"),
        ("nasdaq", "training"),
        ("sp500", "prediction"),
        ("wti", "top10"),
    ],
)
def test_production_source_failure_stops_before_downstream_steps(
    monkeypatch,
    tmp_path: Path,
    missing_source: str,
    skipped_step: str,
) -> None:
    """Strict production source failure should stop before feature/training/prediction/report steps."""
    test_config = _pipeline_test_config(tmp_path)
    _FailingFeatureSourceCompletenessChecker.missing_source = missing_source
    calls = {"feature": 0, "training": 0, "prediction": 0, "top10": 0}

    def forbidden_feature(**_kwargs):
        calls["feature"] += 1
        raise AssertionError("feature update should be skipped")

    def forbidden_training(**_kwargs):
        calls["training"] += 1
        raise AssertionError("training update should be skipped")

    def forbidden_prediction(**_kwargs):
        calls["prediction"] += 1
        raise AssertionError("prediction should be skipped")

    def forbidden_top10(**_kwargs):
        calls["top10"] += 1
        raise AssertionError("Top10 should be skipped")

    monkeypatch.setattr(daily_pipeline, "load_daily_update_config", lambda _path: test_config)
    monkeypatch.setattr(daily_pipeline, "run_ohlcv_update", lambda **_kwargs: _mock_ohlcv_result())
    monkeypatch.setattr(daily_pipeline, "run_macro_update", lambda **_kwargs: _mock_macro_result())
    monkeypatch.setattr(daily_pipeline, "FeatureSourceCompletenessChecker", _FailingFeatureSourceCompletenessChecker)
    monkeypatch.setattr(daily_pipeline, "run_feature_update", forbidden_feature)
    monkeypatch.setattr(daily_pipeline, "run_training_update", forbidden_training)
    monkeypatch.setattr(daily_pipeline, "generate_daily_predictions", forbidden_prediction)
    monkeypatch.setattr(daily_pipeline, "generate_daily_top10_report", forbidden_top10)

    status = daily_pipeline.run_pipeline(
        daily_pipeline.parse_args(["--as-of-date", "2099-04-01", "--skip-download", "--config", str(CONFIG_PATH)])
    )

    assert status.pipeline_stop_reason == "Feature Source Completeness Failed"
    assert status.failed_feature_sources == [missing_source]
    assert status.feature_update_executed is False
    assert status.training_update_executed is False
    assert status.rolling_train_executed is False
    assert status.prediction_executed is False
    assert status.top10_generated is False
    assert calls[skipped_step] == 0


def test_sox_all_null_console_and_status_show_specific_reason(monkeypatch, tmp_path: Path, capsys) -> None:
    """A matching SOX date failure should expose feature diagnostics, not a date contradiction."""
    test_config = _pipeline_test_config(tmp_path)
    _FailingFeatureSourceCompletenessChecker.missing_source = "sox"
    monkeypatch.setattr(daily_pipeline, "load_daily_update_config", lambda _path: test_config)
    monkeypatch.setattr(daily_pipeline, "run_ohlcv_update", lambda **_kwargs: _mock_ohlcv_result())
    monkeypatch.setattr(daily_pipeline, "run_macro_update", lambda **_kwargs: _mock_macro_result())
    monkeypatch.setattr(daily_pipeline, "FeatureSourceCompletenessChecker", _FailingFeatureSourceCompletenessChecker)
    monkeypatch.setattr(daily_pipeline, "run_feature_update", lambda **_kwargs: pytest.fail("feature should be skipped"))

    status = daily_pipeline.run_pipeline(
        daily_pipeline.parse_args(["--as-of-date", "2099-04-01", "--skip-download", "--config", str(CONFIG_PATH)])
    )
    output = capsys.readouterr().out

    assert "SOX......... FAIL" in output
    assert "reason: SOX close exists but sox_return_1d cannot be computed yet because prior SOX history is missing." in output
    assert status.actual_sox_date == status.expected_feature_date
    assert status.sox_check_enabled is True
    assert status.sox_close_present is True
    assert status.sox_return_present is True
    assert status.sox_return_non_null_count == 0
    assert status.sox_failure_reason and "prior SOX history is missing" in status.sox_failure_reason


def test_production_source_failure_creates_summary_and_status(monkeypatch, tmp_path: Path) -> None:
    """Failure should produce markdown summary and status with failed source details."""
    test_config = _pipeline_test_config(tmp_path)
    _FailingFeatureSourceCompletenessChecker.missing_source = "nasdaq"
    monkeypatch.setattr(daily_pipeline, "load_daily_update_config", lambda _path: test_config)
    monkeypatch.setattr(daily_pipeline, "run_ohlcv_update", lambda **_kwargs: _mock_ohlcv_result())
    monkeypatch.setattr(daily_pipeline, "run_macro_update", lambda **_kwargs: _mock_macro_result())
    monkeypatch.setattr(daily_pipeline, "FeatureSourceCompletenessChecker", _FailingFeatureSourceCompletenessChecker)
    monkeypatch.setattr(daily_pipeline, "run_feature_update", lambda **_kwargs: pytest.fail("feature should be skipped"))

    status = daily_pipeline.run_pipeline(
        daily_pipeline.parse_args(["--as-of-date", "2099-04-02", "--skip-download", "--config", str(CONFIG_PATH)])
    )

    status_path = tmp_path / "status" / "daily_update_status_20990402.json"
    assert status_path.exists()
    saved_status = json.loads(status_path.read_text(encoding="utf-8"))
    assert saved_status["pipeline_stop_reason"] == "Feature Source Completeness Failed"
    assert saved_status["failed_feature_sources"] == ["nasdaq"]
    assert saved_status["top10_generated"] is False
    assert status.daily_summary_report is not None
    summary = Path(status.daily_summary_report).read_text(encoding="utf-8")
    assert "Status: FAILED" in summary
    assert "Feature Source Completeness Failed" in summary
    assert "Top10: NOT GENERATED" in summary


def test_main_returns_failure_code_when_strict_source_check_fails(monkeypatch, tmp_path: Path) -> None:
    """CLI main should exit non-zero when production completeness fails."""
    test_config = _pipeline_test_config(tmp_path)
    _FailingFeatureSourceCompletenessChecker.missing_source = "krx"
    monkeypatch.setattr(daily_pipeline, "load_daily_update_config", lambda _path: test_config)
    monkeypatch.setattr(daily_pipeline, "run_ohlcv_update", lambda **_kwargs: _mock_ohlcv_result())
    monkeypatch.setattr(daily_pipeline, "run_macro_update", lambda **_kwargs: _mock_macro_result())
    monkeypatch.setattr(daily_pipeline, "FeatureSourceCompletenessChecker", _FailingFeatureSourceCompletenessChecker)
    monkeypatch.setattr(daily_pipeline, "run_feature_update", lambda **_kwargs: pytest.fail("feature should be skipped"))

    exit_code = daily_pipeline.main(["--as-of-date", "2099-04-03", "--skip-download", "--config", str(CONFIG_PATH)])

    assert exit_code == 1


def test_production_source_all_pass_continues_normal_pipeline(monkeypatch, tmp_path: Path) -> None:
    """Passing completeness should allow the normal downstream mocked pipeline to continue."""
    test_config = _pipeline_test_config(tmp_path)
    monkeypatch.setattr(daily_pipeline, "load_daily_update_config", lambda _path: test_config)
    monkeypatch.setattr(daily_pipeline, "run_ohlcv_update", lambda **_kwargs: _mock_ohlcv_result())
    monkeypatch.setattr(daily_pipeline, "run_macro_update", lambda **_kwargs: _mock_macro_result())
    monkeypatch.setattr(daily_pipeline, "run_feature_update", lambda **_kwargs: _mock_feature_result())
    monkeypatch.setattr(daily_pipeline, "run_training_update", lambda **_kwargs: _mock_training_result())
    monkeypatch.setattr(daily_pipeline, "train_daily_models", lambda **_kwargs: _mock_model_result())
    monkeypatch.setattr(daily_pipeline, "generate_daily_predictions", lambda **_kwargs: _mock_prediction_result())
    monkeypatch.setattr(daily_pipeline, "generate_daily_top10_report", lambda **_kwargs: _mock_report_result())

    status = daily_pipeline.run_pipeline(
        daily_pipeline.parse_args(["--as-of-date", "2099-04-04", "--skip-download", "--config", str(CONFIG_PATH)])
    )

    assert status.pipeline_stop_reason is None
    assert status.feature_source_completeness_passed is True
    assert status.feature_update_executed is True
    assert status.training_update_executed is True
    assert status.rolling_train_executed is True
    assert status.prediction_executed is True
    assert status.top10_generated is True


def make_stale_source_config(tmp_path: Path):
    """Create production config whose feature sources are stale versus target date."""
    config = load_daily_update_config(CONFIG_PATH)
    clean_path = tmp_path / "clean.parquet"
    macro_path = tmp_path / "macro.parquet"
    valid_ohlcv_rows("2026-06-12").to_parquet(clean_path, index=False)
    pd.DataFrame(
        {
            "date": pd.to_datetime(["2026-06-12"]),
            "nasdaq_close": [100.0],
            "sox_close": [200.0],
            "sp500_close": [100.0],
            "vix_close": [20.0],
            "wti_close": [70.0],
            "usdkrw": [1300.0],
        }
    ).to_parquet(macro_path, index=False)
    values = dict(config.values)
    values["clean_ohlcv_file"] = str(clean_path)
    values["macro_file"] = str(macro_path)
    values["feature_file"] = str(tmp_path / "features_not_built.parquet")
    values["daily_status_dir"] = str(tmp_path / "status")
    values["log_dir"] = str(tmp_path / "logs")
    values["daily_report_dir"] = str(tmp_path / "reports")
    values["production_mode"] = True
    values["strict_feature_source_check"] = True
    values["enable_us10y_check"] = False
    values["enable_gold_check"] = False
    values["enable_dxy_check"] = False
    return type(config)(config_path=config.config_path, values=values, project_root=PROJECT_ROOT)


def test_strict_production_stale_fallback_blocks_prediction_and_top10(monkeypatch, tmp_path: Path, capsys) -> None:
    """Strict production must validate target_update_date and stop on stale fallback data."""
    test_config = make_stale_source_config(tmp_path)
    monkeypatch.setattr(daily_pipeline, "load_daily_update_config", lambda _path: test_config)
    monkeypatch.setattr(daily_pipeline, "run_ohlcv_update", lambda **_kwargs: _mock_fallback_ohlcv_result())
    monkeypatch.setattr(daily_pipeline, "run_macro_update", lambda **_kwargs: _mock_macro_result())
    monkeypatch.setattr(daily_pipeline, "FeatureSourceCompletenessChecker", FeatureSourceCompletenessChecker)
    monkeypatch.setattr(daily_pipeline, "run_feature_update", lambda **_kwargs: pytest.fail("feature should be skipped"))
    monkeypatch.setattr(daily_pipeline, "run_training_update", lambda **_kwargs: pytest.fail("training should be skipped"))
    monkeypatch.setattr(daily_pipeline, "generate_daily_predictions", lambda **_kwargs: pytest.fail("prediction should be skipped"))
    monkeypatch.setattr(daily_pipeline, "generate_daily_top10_report", lambda **_kwargs: pytest.fail("top10 should be skipped"))

    status = daily_pipeline.run_pipeline(
        daily_pipeline.parse_args(["--as-of-date", "2026-06-17", "--skip-download", "--config", str(CONFIG_PATH)])
    )

    assert status.target_update_date == "2026-06-16"
    assert status.update_date == "2026-06-12"
    assert status.expected_feature_date == "2026-06-16"
    assert status.feature_source_completeness_passed is False
    assert status.pipeline_stop_reason == "Feature Source Completeness Failed: stale data"
    assert status.stale_data_blocked_by_production_policy is True
    assert status.failed_feature_sources == ["krx"]
    assert status.feature_update_executed is False
    assert status.training_update_executed is False
    assert status.rolling_train_executed is False
    assert status.prediction_executed is False
    assert status.top10_generated is False
    assert status.prediction_output_csv is None
    assert status.top10_report_csv is None
    assert status.daily_summary_report is not None
    summary = Path(status.daily_summary_report).read_text(encoding="utf-8")
    assert "Production stopped because latest available feature source data is stale." in summary
    assert "Target Update Date: 2026-06-16" in summary
    assert "Latest Clean Data Date: 2026-06-12" in summary
    output = capsys.readouterr().out
    assert "AI Trading Daily Update Pipeline - STOPPED" in output
    assert "STRICT_PRODUCTION_STOP_EXIT_CODE=1" in output
    assert "Latest required market data is unavailable." in output
    assert "Target update date:" in output
    assert "Latest available data:" in output
    assert "Top10:" in output
    assert "NOT GENERATED" in output
    assert "Prediction:" in output
    assert "NOT EXECUTED" in output
    assert "Exit code:" in output
    assert "Downloaded market date" in output
    assert status.pipeline_exit_code == 1
    assert status.pipeline_exit_message == "Latest required market data unavailable"
    assert status.latest_available_market_date == "2026-06-12"
    assert int(status) == 1


def test_strict_production_stale_fallback_main_returns_failure(monkeypatch, tmp_path: Path) -> None:
    """Strict stale-data stop should produce a non-zero CLI exit code."""
    test_config = make_stale_source_config(tmp_path)
    monkeypatch.setattr(daily_pipeline, "load_daily_update_config", lambda _path: test_config)
    monkeypatch.setattr(daily_pipeline, "run_ohlcv_update", lambda **_kwargs: _mock_fallback_ohlcv_result())
    monkeypatch.setattr(daily_pipeline, "run_macro_update", lambda **_kwargs: _mock_macro_result())
    monkeypatch.setattr(daily_pipeline, "FeatureSourceCompletenessChecker", FeatureSourceCompletenessChecker)
    monkeypatch.setattr(daily_pipeline, "run_feature_update", lambda **_kwargs: pytest.fail("feature should be skipped"))
    monkeypatch.setattr(daily_pipeline.time, "sleep", lambda _seconds: pytest.fail("production stop should not sleep"))

    exit_code = daily_pipeline.main(["--as-of-date", "2026-06-17", "--skip-download", "--config", str(CONFIG_PATH)])

    assert exit_code == 1


def test_check_sources_only_stops_after_source_check(monkeypatch, tmp_path: Path) -> None:
    """--check-sources-only should write status and skip downstream pipeline work."""
    test_config = _pipeline_test_config(tmp_path)
    monkeypatch.setattr(daily_pipeline, "load_daily_update_config", lambda _path: test_config)
    monkeypatch.setattr(daily_pipeline, "run_ohlcv_update", lambda **_kwargs: _mock_ohlcv_result())
    monkeypatch.setattr(daily_pipeline, "run_macro_update", lambda **_kwargs: _mock_macro_result())
    monkeypatch.setattr(daily_pipeline, "run_feature_update", lambda **_kwargs: pytest.fail("feature should be skipped"))

    status = daily_pipeline.run_pipeline(
        daily_pipeline.parse_args(["--as-of-date", "2099-04-05", "--skip-download", "--check-sources-only", "--config", str(CONFIG_PATH)])
    )

    assert int(status) == 0
    assert status.pipeline_exit_message == "Source completeness check passed"
    assert "status_written" in status.completed_steps
    assert status.feature_update_executed is False


class FakePykrxStock:
    """Small fake pykrx stock module for diagnostics tests."""

    def get_market_ohlcv_by_ticker(self, _date: str, market: str) -> pd.DataFrame:
        return pd.DataFrame({"ticker": ["005930", "000660"], "market": [market, market], "close": [100.0, 101.0]})

    def get_market_ticker_list(self, _date: str, market: str) -> list[str]:
        return ["005930", "000660"] if market == "KOSPI" else ["035720"]

    def get_market_ohlcv_by_date(self, _start: str, _end: str, ticker: str) -> pd.DataFrame:
        return pd.DataFrame({"ticker": [ticker], "close": [100.0]})


def test_pykrx_diagnostic_market_wide_endpoint_works() -> None:
    """Diagnostic should exercise market-wide, ticker-list, and sample ticker endpoints."""
    results = diagnose_download.run_diagnostics(
        "2026-06-16",
        sample_size=2,
        timeout_seconds=1,
        stock_module=FakePykrxStock(),
    )

    endpoints = [result.endpoint for result in results]
    assert "stock.get_market_ohlcv_by_ticker KOSPI" in endpoints
    assert "stock.get_market_ohlcv_by_ticker KOSDAQ" in endpoints
    assert "stock.get_market_ticker_list KOSPI" in endpoints
    assert "stock.get_market_ticker_list KOSDAQ" in endpoints
    assert "stock.get_market_ohlcv_by_date 005930" in endpoints
    assert "20240614" in {result.date_input for result in results}
    assert "2024-06-14" in {result.date_input for result in results}
    assert "20260616" in {result.date_input for result in results}
    assert all(result.success for result in results)


class EmptyPykrxStock(FakePykrxStock):
    """Fake pykrx stock module returning no rows."""

    def get_market_ohlcv_by_ticker(self, _date: str, market: str) -> pd.DataFrame:
        return pd.DataFrame()

    def get_market_ticker_list(self, _date: str, market: str) -> list[str]:
        return []

    def get_market_ohlcv_by_date(self, _start: str, _end: str, ticker: str) -> pd.DataFrame:
        return pd.DataFrame()


class MissingColumnsPykrxStock(FakePykrxStock):
    """Fake pykrx stock module returning unexpected OHLCV columns."""

    def get_market_ohlcv_by_ticker(self, _date: str, market: str) -> pd.DataFrame:
        return pd.DataFrame({"unexpected": [1]})

    def get_market_ohlcv_by_date(self, _start: str, _end: str, ticker: str) -> pd.DataFrame:
        return pd.DataFrame({"unexpected": [ticker]})


def test_pykrx_diagnostic_empty_ticker_list_is_unavailable() -> None:
    """Empty pykrx responses should be reported as unavailable data."""
    results = diagnose_download.run_diagnostics(
        "2026-06-16",
        sample_size=2,
        timeout_seconds=1,
        stock_module=EmptyPykrxStock(),
    )

    assert any(result.endpoint == "stock.get_market_ticker_list KOSPI" for result in results)
    assert all(result.success is False for result in results)
    assert all("endpoint returned empty" in result.exception_message for result in results)


def test_pykrx_diagnostic_missing_columns_is_unavailable(capsys) -> None:
    """Unexpected OHLCV payloads should be reported as unavailable, not as login errors."""
    results = diagnose_download.run_diagnostics(
        "2026-06-16",
        sample_size=1,
        timeout_seconds=1,
        stock_module=MissingColumnsPykrxStock(),
    )

    diagnose_download.print_results(results)
    output = capsys.readouterr().out
    assert any("missing expected columns" in result.exception_message for result in results)
    assert "likely market data unavailable for date" in output
    assert "KRX 로그인 실패" not in output
    assert "KRX_ID" not in output
    assert "KRX_PW" not in output


def test_pykrx_diagnostic_writes_markdown_report(tmp_path: Path) -> None:
    """Diagnostic report should include environment, network, and endpoint matrix."""
    environment = {
        "python_version": "test-python",
        "pykrx_version": "test-pykrx",
        "pandas_version": "test-pandas",
        "requests_version": "test-requests",
        "current_working_directory": str(PROJECT_ROOT),
        "system_time": "2099-01-01T00:00:00",
    }
    network = diagnose_download.NetworkCheckResult(
        elapsed_seconds=0.01,
        success=True,
        status_code=200,
        content_type="text/html",
        preview="<html>",
        exception_type="",
        exception_message="",
    )
    results = [
        diagnose_download.DiagnosticResult(
            endpoint="stock.get_market_ticker_list KOSPI",
            elapsed_seconds=0.01,
            success=True,
            rows_returned=1,
            exception_message="",
            date_input="20260616",
            payload_type="list",
            shape="(1,)",
            preview_rows=[{"value": "005930"}],
        )
    ]

    report_path = diagnose_download.write_report("2026-06-16", environment, network, results, report_dir=tmp_path)
    report_text = report_path.read_text(encoding="utf-8")

    assert report_path.name == "pykrx_diagnostic_20260616.md"
    assert "## Environment" in report_text
    assert "## Network Sanity Check" in report_text
    assert "## Endpoint Matrix" in report_text
    assert "stock.get_market_ticker_list KOSPI" in report_text


def test_pykrx_diagnostic_recommends_per_ticker_when_only_per_ticker_works() -> None:
    """Diagnostic should recommend per-ticker endpoint when market-wide endpoints fail."""
    results = [
        diagnose_download.DiagnosticResult("stock.get_market_ohlcv_by_ticker KOSPI", 0.01, False, 0, "missing"),
        diagnose_download.DiagnosticResult("stock.get_market_ticker_list KOSPI", 0.01, False, 0, "empty"),
        diagnose_download.DiagnosticResult("stock.get_market_ohlcv_by_date 005930", 0.01, True, 1, ""),
    ]

    assert diagnose_download.final_recommendation(results) == "Use per-ticker endpoint"


def test_pykrx_diagnostic_timeout_is_reported_without_hanging() -> None:
    """Diagnostic endpoint timeout should return TIMEOUT and continue."""
    started = threading.Event()
    release = threading.Event()

    def slow_call() -> pd.DataFrame:
        started.set()
        release.wait(1.0)
        return pd.DataFrame({"x": [1]})

    result = diagnose_download.diagnose_endpoint("slow endpoint", slow_call, timeout_seconds=0.01)
    release.set()

    assert started.is_set()
    assert result.success is False
    assert result.exception_message == "TIMEOUT"
    assert result.rows_returned == 0


def test_diagnose_download_cli_does_not_train_or_predict(monkeypatch, tmp_path: Path) -> None:
    """--diagnose-download should stop before feature/model/prediction/report steps."""
    test_config = _pipeline_test_config(tmp_path)
    monkeypatch.setattr(daily_pipeline, "load_daily_update_config", lambda _path: test_config)
    monkeypatch.setattr(
        "scripts.diagnose_pykrx_daily_download.run_diagnostics",
        lambda _date: [diagnose_download.DiagnosticResult("fake", 0.0, True, 1, "")],
    )
    monkeypatch.setattr(daily_pipeline, "run_feature_update", lambda **_kwargs: pytest.fail("feature should be skipped"))
    monkeypatch.setattr(daily_pipeline, "train_daily_models", lambda **_kwargs: pytest.fail("training should be skipped"))
    monkeypatch.setattr(daily_pipeline, "generate_daily_predictions", lambda **_kwargs: pytest.fail("prediction should be skipped"))
    monkeypatch.setattr(daily_pipeline, "generate_daily_top10_report", lambda **_kwargs: pytest.fail("top10 should be skipped"))

    status = daily_pipeline.run_pipeline(
        daily_pipeline.parse_args(["--as-of-date", "2099-04-06", "--diagnose-download", "--config", str(CONFIG_PATH)])
    )

    assert int(status) == 0
    assert status.pipeline_exit_message == "Download diagnostics completed"
    assert "download_diagnostics_completed" in status.completed_steps
    assert status.prediction_executed is False
    assert status.top10_generated is False


def test_ohlcv_validation_accepts_valid_rows() -> None:
    """Valid OHLCV rows should pass into clean set."""
    valid, invalid = split_valid_invalid_ohlcv(valid_ohlcv_rows())

    assert len(valid) == 2
    assert invalid.empty


def test_ohlcv_validation_rejects_zero_prices() -> None:
    """Zero open/high/low/close should be invalid."""
    rows = valid_ohlcv_rows()
    rows.loc[0, "open"] = 0
    rows.loc[1, "close"] = 0

    valid, invalid = split_valid_invalid_ohlcv(rows)

    assert valid.empty
    assert set(invalid["ticker"]) == {"005930", "000660"}


def test_ohlcv_validation_rejects_invalid_high_low_relations() -> None:
    """Invalid high/low relations should be excluded."""
    rows = valid_ohlcv_rows()
    rows.loc[0, "high"] = 80
    rows.loc[1, "low"] = 230

    valid, invalid = split_valid_invalid_ohlcv(rows)

    assert valid.empty
    assert len(invalid) == 2


def test_safe_append_avoids_duplicate_date_ticker(tmp_path: Path) -> None:
    """Safe append should not duplicate rows already present."""
    path = tmp_path / "ohlcv.parquet"
    existing = valid_ohlcv_rows()
    existing.to_parquet(path, index=False)

    combined, added = safe_append_ohlcv(path, valid_ohlcv_rows(), pd.Timestamp("2026-06-12"), force=False)

    assert added == 0
    assert len(combined) == 2
    assert not combined.duplicated(subset=["date", "ticker"]).any()


def test_force_replace_replaces_same_date_rows(tmp_path: Path) -> None:
    """Force should replace same-date rows while preserving no duplicates."""
    path = tmp_path / "ohlcv.parquet"
    existing = valid_ohlcv_rows()
    existing.to_parquet(path, index=False)
    replacement = valid_ohlcv_rows()
    replacement["close"] = [111.0, 222.0]

    combined, added = safe_append_ohlcv(path, replacement, pd.Timestamp("2026-06-12"), force=True)

    assert added == 2
    assert sorted(combined["close"].tolist()) == [111.0, 222.0]
    assert not combined.duplicated(subset=["date", "ticker"]).any()


def test_dry_run_does_not_write_files(tmp_path: Path) -> None:
    """Dry-run should not write raw/clean files or snapshots."""
    config = load_daily_update_config(CONFIG_PATH)
    values = dict(config.values)
    values["raw_ohlcv_file"] = str(tmp_path / "raw.parquet")
    values["clean_ohlcv_file"] = str(tmp_path / "clean.parquet")
    values["daily_raw_dir"] = str(tmp_path / "daily_raw")
    values["daily_processed_dir"] = str(tmp_path / "daily_processed")
    test_config = type(config)(config_path=config.config_path, values=values, project_root=PROJECT_ROOT)
    context = build_daily_run_context(test_config, "2026-06-12", True, True, False)

    result = run_ohlcv_update(test_config, context, dry_run=True, skip_download=True, force=False)

    assert result.raw_rows_added == 0
    assert not (tmp_path / "raw.parquet").exists()
    assert not (tmp_path / "clean.parquet").exists()
    assert not (tmp_path / "daily_raw").exists()


def test_skip_download_does_not_call_pykrx(tmp_path: Path) -> None:
    """Skip-download should not call the downloader function."""
    config = load_daily_update_config(CONFIG_PATH)
    raw_path = tmp_path / "raw.parquet"
    clean_path = tmp_path / "clean.parquet"
    values = dict(config.values)
    values["raw_ohlcv_file"] = str(raw_path)
    values["clean_ohlcv_file"] = str(clean_path)
    test_config = type(config)(config_path=config.config_path, values=values, project_root=PROJECT_ROOT)
    rows = valid_ohlcv_rows("2026-06-12")
    rows.to_parquet(raw_path, index=False)
    context = build_daily_run_context(test_config, "2026-06-15", False, True, False)

    def forbidden_downloader(_tickers: list[str], _date: pd.Timestamp) -> pd.DataFrame:
        raise AssertionError("Downloader should not be called")

    result = run_ohlcv_update(
        test_config,
        context,
        dry_run=False,
        skip_download=True,
        force=False,
        downloader=forbidden_downloader,
    )

    assert result.raw_rows_downloaded_or_found == 2
    assert result.ohlcv_download_mode == "no_download"
    assert result.ohlcv_download_attempts == 0


def test_no_download_skips_krx_call(monkeypatch, tmp_path: Path) -> None:
    """--no-download should skip KRX while still completing downstream mocked steps."""
    as_of_date = "2099-03-01"
    config = load_daily_update_config(CONFIG_PATH)
    values = dict(config.values)
    values["daily_status_dir"] = str(tmp_path / "status")
    values["log_dir"] = str(tmp_path / "logs")
    test_config = type(config)(config_path=config.config_path, values=values, project_root=PROJECT_ROOT)
    status_file = tmp_path / "status" / "daily_update_status_20990301.json"
    monkeypatch.setattr(daily_pipeline, "load_daily_update_config", lambda _path: test_config)
    monkeypatch.setattr(daily_pipeline, "run_ohlcv_update", lambda **kwargs: _mock_ohlcv_result())
    monkeypatch.setattr(daily_pipeline, "run_macro_update", lambda **_kwargs: _mock_macro_result())
    monkeypatch.setattr(daily_pipeline, "run_feature_update", lambda **_kwargs: _mock_feature_result())
    monkeypatch.setattr(daily_pipeline, "run_training_update", lambda **_kwargs: _mock_training_result())
    monkeypatch.setattr(daily_pipeline, "train_daily_models", lambda **_kwargs: _mock_model_result())
    monkeypatch.setattr(daily_pipeline, "generate_daily_predictions", lambda **_kwargs: _mock_prediction_result())
    monkeypatch.setattr(daily_pipeline, "generate_daily_top10_report", lambda **_kwargs: _mock_report_result())

    daily_pipeline.run_pipeline(
        daily_pipeline.parse_args(["--as-of-date", as_of_date, "--no-download", "--config", str(CONFIG_PATH)])
    )

    status = json.loads(status_file.read_text(encoding="utf-8"))
    assert status["skip_download"] is True
    assert "daily_predictions_generated" in status["completed_steps"]
    assert "top10_report_generated" in status["completed_steps"]


def test_pykrx_unavailable_falls_back_and_records_status(monkeypatch, tmp_path: Path, capsys) -> None:
    """pykrx unavailable status should fallback and continue in non-production mode."""
    config = load_daily_update_config(CONFIG_PATH)
    clean_path = tmp_path / "clean.parquet"
    valid_ohlcv_rows("2026-06-12").to_parquet(clean_path, index=False)
    values = dict(config.values)
    values["daily_status_dir"] = str(tmp_path / "status")
    values["log_dir"] = str(tmp_path / "logs")
    values["clean_ohlcv_file"] = str(clean_path)
    values["production_mode"] = False
    test_config = type(config)(config_path=config.config_path, values=values, project_root=PROJECT_ROOT)

    def fake_ohlcv(**kwargs):
        return _mock_fallback_ohlcv_result()

    monkeypatch.setattr(daily_pipeline, "load_daily_update_config", lambda _path: test_config)
    monkeypatch.setattr(daily_pipeline, "run_ohlcv_update", fake_ohlcv)
    monkeypatch.setattr(daily_pipeline, "run_macro_update", lambda **_kwargs: _mock_macro_result())
    monkeypatch.setattr(daily_pipeline, "run_feature_update", lambda **_kwargs: _mock_feature_result())
    monkeypatch.setattr(daily_pipeline, "run_training_update", lambda **_kwargs: _mock_training_result())
    monkeypatch.setattr(daily_pipeline, "train_daily_models", lambda **_kwargs: _mock_model_result())
    monkeypatch.setattr(daily_pipeline, "generate_daily_predictions", lambda **_kwargs: _mock_prediction_result())
    monkeypatch.setattr(daily_pipeline, "generate_daily_top10_report", lambda **_kwargs: _mock_report_result())

    daily_pipeline.run_pipeline(daily_pipeline.parse_args(["--as-of-date", "2026-06-16", "--config", str(CONFIG_PATH)]))

    output = capsys.readouterr().out
    status = json.loads((tmp_path / "status" / "daily_update_status_20260616.json").read_text(encoding="utf-8"))
    assert "KRX data for target date 2026-06-15 is unavailable." in output
    assert status["pykrx_data_unavailable"] is True
    assert status["pykrx_empty_response"] is True
    assert status["pykrx_rows_returned"] == 0
    assert status["used_existing_data_fallback"] is True
    assert "daily_predictions_generated" in status["completed_steps"]
    assert "top10_report_generated" in status["completed_steps"]


def test_pipeline_no_longer_imports_krx_login() -> None:
    """Daily pipeline should not import or call authenticated KRX login helpers."""
    source = SCRIPT_PATH.read_text(encoding="utf-8")

    assert "src.pipeline.krx_login" not in source
    assert "attempt_krx_login" not in source


def test_no_local_pykrx_shadowing() -> None:
    """Project should not shadow the installed pykrx package."""
    assert not (PROJECT_ROOT / "pykrx.py").exists()
    assert not (PROJECT_ROOT / "pykrx").exists()


def test_importing_pykrx_has_no_krx_login_output() -> None:
    """Importing pykrx should not print KRX login or credential messages."""
    command = "import pykrx, sys; print(sys.version); print(pykrx.__version__)"
    result = subprocess.run(
        [pykrx_test_python(), "-c", command],
        cwd=PROJECT_ROOT,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        timeout=30,
        check=False,
    )
    combined = result.stdout + result.stderr

    assert result.returncode == 0
    assert "KRX 로그인" not in combined
    assert "KRX_ID" not in combined
    assert "KRX_PW" not in combined
    assert result.stderr == ""
    assert len(result.stdout.strip().splitlines()) == 2


def test_importing_diagnose_script_has_no_krx_login_output() -> None:
    """Importing the diagnostic module should not trigger pykrx login output."""
    command = "import scripts.diagnose_pykrx_daily_download; print('imported')"
    result = subprocess.run(
        [pykrx_test_python(), "-c", command],
        cwd=PROJECT_ROOT,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        timeout=30,
        check=False,
    )
    combined = result.stdout + result.stderr

    assert result.returncode == 0
    assert result.stdout.strip() == "imported"
    assert "KRX 로그인" not in combined
    assert "KRX_ID" not in combined
    assert "KRX_PW" not in combined
    assert result.stderr == ""


def test_pykrx_missing_columns_falls_back(monkeypatch, tmp_path: Path) -> None:
    """pykrx missing-column status should be recorded without crashing."""
    config = load_daily_update_config(CONFIG_PATH)
    clean_path = tmp_path / "clean.parquet"
    valid_ohlcv_rows("2026-06-12").to_parquet(clean_path, index=False)
    values = dict(config.values)
    values["daily_status_dir"] = str(tmp_path / "status")
    values["log_dir"] = str(tmp_path / "logs")
    values["clean_ohlcv_file"] = str(clean_path)
    values["production_mode"] = False
    test_config = type(config)(config_path=config.config_path, values=values, project_root=PROJECT_ROOT)
    failed_ohlcv = SimpleNamespace(
        **{
            **_mock_fallback_ohlcv_result().__dict__,
            "pykrx_empty_response": False,
            "pykrx_missing_columns": True,
            "warnings": ["pykrx_missing_columns"],
        }
    )
    monkeypatch.setattr(daily_pipeline, "load_daily_update_config", lambda _path: test_config)
    monkeypatch.setattr(daily_pipeline, "run_ohlcv_update", lambda **_kwargs: failed_ohlcv)
    monkeypatch.setattr(daily_pipeline, "run_macro_update", lambda **_kwargs: _mock_macro_result())
    monkeypatch.setattr(daily_pipeline, "run_feature_update", lambda **_kwargs: _mock_feature_result())
    monkeypatch.setattr(daily_pipeline, "run_training_update", lambda **_kwargs: _mock_training_result())
    monkeypatch.setattr(daily_pipeline, "train_daily_models", lambda **_kwargs: _mock_model_result())
    monkeypatch.setattr(daily_pipeline, "generate_daily_predictions", lambda **_kwargs: _mock_prediction_result())
    monkeypatch.setattr(daily_pipeline, "generate_daily_top10_report", lambda **_kwargs: _mock_report_result())

    daily_pipeline.run_pipeline(daily_pipeline.parse_args(["--as-of-date", "2026-06-16", "--config", str(CONFIG_PATH)]))

    status = json.loads((tmp_path / "status" / "daily_update_status_20260616.json").read_text(encoding="utf-8"))
    assert status["pykrx_missing_columns"] is True
    assert status["pykrx_data_unavailable"] is True
    assert status["used_existing_data_fallback"] is True


def test_no_login_is_compatibility_noop(monkeypatch, tmp_path: Path) -> None:
    """--no-login should not force pykrx download skipping in the pykrx-only path."""
    config = load_daily_update_config(CONFIG_PATH)
    values = dict(config.values)
    values["daily_status_dir"] = str(tmp_path / "status")
    values["log_dir"] = str(tmp_path / "logs")
    test_config = type(config)(config_path=config.config_path, values=values, project_root=PROJECT_ROOT)

    def fake_ohlcv(**kwargs):
        assert kwargs["skip_download"] is False
        return _mock_ohlcv_result()

    monkeypatch.setattr(daily_pipeline, "load_daily_update_config", lambda _path: test_config)
    monkeypatch.setattr(daily_pipeline, "run_ohlcv_update", fake_ohlcv)
    monkeypatch.setattr(daily_pipeline, "run_macro_update", lambda **_kwargs: _mock_macro_result())
    monkeypatch.setattr(daily_pipeline, "run_feature_update", lambda **_kwargs: _mock_feature_result())
    monkeypatch.setattr(daily_pipeline, "run_training_update", lambda **_kwargs: _mock_training_result())
    monkeypatch.setattr(daily_pipeline, "train_daily_models", lambda **_kwargs: _mock_model_result())
    monkeypatch.setattr(daily_pipeline, "generate_daily_predictions", lambda **_kwargs: _mock_prediction_result())
    monkeypatch.setattr(daily_pipeline, "generate_daily_top10_report", lambda **_kwargs: _mock_report_result())

    daily_pipeline.run_pipeline(
        daily_pipeline.parse_args(["--as-of-date", "2099-03-03", "--no-login", "--config", str(CONFIG_PATH)])
    )

    status = json.loads((tmp_path / "status" / "daily_update_status_20990303.json").read_text(encoding="utf-8"))
    assert status["skip_download"] is False
    assert status["krx_login_attempted"] is False
    assert status["krx_login_success"] is False


def test_successful_download_updates_prediction_date_to_as_of_date(monkeypatch, tmp_path: Path) -> None:
    """Successful target-date download should keep target update and next prediction date."""
    config = load_daily_update_config(CONFIG_PATH)
    clean_path = tmp_path / "clean.parquet"
    valid_ohlcv_rows("2026-06-12").to_parquet(clean_path, index=False)
    values = dict(config.values)
    values["daily_status_dir"] = str(tmp_path / "status")
    values["log_dir"] = str(tmp_path / "logs")
    values["clean_ohlcv_file"] = str(clean_path)
    test_config = type(config)(config_path=config.config_path, values=values, project_root=PROJECT_ROOT)
    monkeypatch.setattr(daily_pipeline, "load_daily_update_config", lambda _path: test_config)
    monkeypatch.setattr(daily_pipeline, "run_ohlcv_update", lambda **_kwargs: _mock_ohlcv_result())
    monkeypatch.setattr(daily_pipeline, "run_macro_update", lambda **_kwargs: _mock_macro_result())
    monkeypatch.setattr(daily_pipeline, "run_feature_update", lambda **_kwargs: _mock_feature_result())
    monkeypatch.setattr(daily_pipeline, "run_training_update", lambda **_kwargs: _mock_training_result())
    monkeypatch.setattr(daily_pipeline, "train_daily_models", lambda **_kwargs: _mock_model_result())
    monkeypatch.setattr(daily_pipeline, "generate_daily_predictions", lambda **_kwargs: _mock_prediction_result())
    monkeypatch.setattr(daily_pipeline, "generate_daily_top10_report", lambda **_kwargs: _mock_report_result())

    daily_pipeline.run_pipeline(
        daily_pipeline.parse_args(["--as-of-date", "2026-06-16", "--config", str(CONFIG_PATH)])
    )

    status = json.loads((tmp_path / "status" / "daily_update_status_20260616.json").read_text(encoding="utf-8"))
    assert status["target_update_date"] == "2026-06-15"
    assert status["attempted_download_date"] == "2026-06-15"
    assert status["downloaded_update_date"] == "2026-06-15"
    assert status["update_date"] == "2026-06-15"
    assert status["prediction_date"] == "2026-06-16"
    assert status["old_data_warning"] is None


def test_failed_download_status_falls_back_with_old_data_warning(monkeypatch, tmp_path: Path) -> None:
    """Failed target-date download should fallback to latest clean data and warn OLD_DATA in non-production mode."""
    config = load_daily_update_config(CONFIG_PATH)
    clean_path = tmp_path / "clean.parquet"
    valid_ohlcv_rows("2026-06-12").to_parquet(clean_path, index=False)
    values = dict(config.values)
    values["daily_status_dir"] = str(tmp_path / "status")
    values["log_dir"] = str(tmp_path / "logs")
    values["clean_ohlcv_file"] = str(clean_path)
    values["production_mode"] = False
    test_config = type(config)(config_path=config.config_path, values=values, project_root=PROJECT_ROOT)
    failed_ohlcv = _mock_ohlcv_result()
    failed_ohlcv = SimpleNamespace(
        **{
            **failed_ohlcv.__dict__,
            "raw_rows_downloaded_or_found": 0,
            "ohlcv_download_mode": "fallback_existing",
            "ohlcv_download_attempts": 2,
            "ohlcv_download_failed": True,
            "ohlcv_download_error": "download_timed_out_after_30s",
            "used_existing_data_fallback": True,
            "warnings": ["ohlcv_download_fallback_using_existing_clean_data_for_2026-06-15"],
        }
    )
    monkeypatch.setattr(daily_pipeline, "load_daily_update_config", lambda _path: test_config)
    monkeypatch.setattr(daily_pipeline, "run_ohlcv_update", lambda **_kwargs: failed_ohlcv)
    monkeypatch.setattr(daily_pipeline, "run_macro_update", lambda **_kwargs: _mock_macro_result())
    monkeypatch.setattr(daily_pipeline, "run_feature_update", lambda **_kwargs: _mock_feature_result())
    monkeypatch.setattr(daily_pipeline, "run_training_update", lambda **_kwargs: _mock_training_result())
    monkeypatch.setattr(daily_pipeline, "train_daily_models", lambda **_kwargs: _mock_model_result())
    monkeypatch.setattr(daily_pipeline, "generate_daily_predictions", lambda **_kwargs: _mock_prediction_result())
    monkeypatch.setattr(daily_pipeline, "generate_daily_top10_report", lambda **_kwargs: _mock_report_result())

    daily_pipeline.run_pipeline(
        daily_pipeline.parse_args(["--as-of-date", "2026-06-16", "--config", str(CONFIG_PATH)])
    )

    status = json.loads((tmp_path / "status" / "daily_update_status_20260616.json").read_text(encoding="utf-8"))
    assert status["target_update_date"] == "2026-06-15"
    assert status["update_date"] == "2026-06-12"
    assert status["prediction_date"] == "2026-06-15"
    assert status["used_existing_data_fallback"] is True
    assert "OLD_DATA" in status["old_data_warning"]


def test_download_timeout_falls_back() -> None:
    """A downloader timeout should return a fallback outcome."""
    def slow_downloader(_tickers: list[str], _date: pd.Timestamp) -> pd.DataFrame:
        import time

        time.sleep(0.2)
        return valid_ohlcv_rows()

    outcome = download_with_retries(
        slow_downloader,
        ["005930"],
        pd.Timestamp("2026-06-12"),
        timeout_seconds=0.01,
        max_retries=1,
        retry_sleep_seconds=0,
    )

    assert outcome.failed is True
    assert outcome.timed_out is True
    assert outcome.mode == "fallback_existing"


def test_timeout_worker_is_daemon_and_does_not_leave_non_daemon_thread() -> None:
    """Timeout wrapper should not leave a non-daemon worker capable of blocking process exit."""
    started = threading.Event()
    release = threading.Event()
    before = {thread.ident for thread in threading.enumerate() if not thread.daemon}

    def stuck_downloader(_tickers: list[str], _date: pd.Timestamp) -> pd.DataFrame:
        started.set()
        release.wait(1.0)
        return valid_ohlcv_rows()

    with pytest.raises(TimeoutError):
        call_with_timeout(stuck_downloader, ["005930"], pd.Timestamp("2026-06-12"), timeout_seconds=0.01)

    assert started.is_set()
    after = {thread.ident for thread in threading.enumerate() if not thread.daemon}
    release.set()
    assert after <= before | {threading.current_thread().ident}


def test_failed_download_falls_back() -> None:
    """A downloader exception should be captured and converted to fallback."""
    def failing_downloader(_tickers: list[str], _date: pd.Timestamp) -> pd.DataFrame:
        raise RuntimeError("network unavailable")

    outcome = download_with_retries(
        failing_downloader,
        ["005930"],
        pd.Timestamp("2026-06-12"),
        timeout_seconds=1,
        max_retries=2,
        retry_sleep_seconds=0,
    )

    assert outcome.failed is True
    assert outcome.attempts == 2
    assert outcome.mode == "fallback_existing"
    assert "network unavailable" in str(outcome.error)


def test_download_missing_pykrx_columns_is_unavailable() -> None:
    """pykrx missing-column errors should be treated as data unavailable."""
    def bad_downloader(_tickers: list[str], _date: pd.Timestamp) -> pd.DataFrame:
        raise ValueError("pykrx_missing_columns:['date', 'close']")

    outcome = download_with_retries(
        bad_downloader,
        ["005930"],
        pd.Timestamp("2026-06-12"),
        timeout_seconds=1,
        max_retries=2,
        retry_sleep_seconds=0,
    )

    assert outcome.failed is True
    assert outcome.missing_columns is True
    assert outcome.data_unavailable is True
    assert outcome.attempts == 1


class FakePerTickerStock:
    """Fake pykrx stock module for per-ticker OHLCV download tests."""

    def __init__(self, missing_trading_value: bool = False, empty_tickers: set[str] | None = None):
        self.missing_trading_value = missing_trading_value
        self.empty_tickers = empty_tickers or set()
        self.calls: list[tuple[str, str, str]] = []

    def get_market_ohlcv_by_date(self, start: str, end: str, ticker: str) -> pd.DataFrame:
        self.calls.append((start, end, ticker))
        if ticker in self.empty_tickers:
            return pd.DataFrame()
        data = {
            "시가": [100.0],
            "고가": [110.0],
            "저가": [90.0],
            "종가": [105.0],
            "거래량": [10.0],
        }
        if not self.missing_trading_value:
            data["거래대금"] = [1050.0]
        return pd.DataFrame(data, index=pd.to_datetime([start]))


def test_per_ticker_download_creates_normalized_ohlcv_rows() -> None:
    """Per-ticker downloader should normalize one target-date row per ticker."""
    stock = FakePerTickerStock()

    result = download_daily_ohlcv(
        ["005930", "000660"],
        pd.Timestamp("2026-06-16"),
        timeout_seconds=1,
        retry_sleep_seconds=0,
        stock_module=stock,
    )

    assert list(result.columns) == ["date", "ticker", "open", "high", "low", "close", "volume", "trading_value"]
    assert result["ticker"].tolist() == ["005930", "000660"]
    assert result["date"].dt.strftime("%Y-%m-%d").unique().tolist() == ["2026-06-16"]
    assert all(start == end == "20260616" for start, end, _ticker in stock.calls)
    assert result.attrs["tickers_requested"] == 2
    assert result.attrs["tickers_downloaded"] == 2
    assert result.attrs["success_ratio"] == 1.0
    assert result.attrs["enforce_success_threshold"] is True


def test_per_ticker_missing_trading_value_estimates_trading_value() -> None:
    """Missing 거래대금 should be estimated from close * volume and counted."""
    stock = FakePerTickerStock(missing_trading_value=True)

    result = download_daily_ohlcv(
        ["005930"],
        pd.Timestamp("2026-06-16"),
        timeout_seconds=1,
        retry_sleep_seconds=0,
        stock_module=stock,
    )

    assert result.loc[0, "trading_value"] == 1050.0
    assert result.attrs["trading_value_estimated_count"] == 1


def test_low_success_ratio_fails_production_without_appending_partial_data(tmp_path: Path) -> None:
    """Strict production should reject and not append low-coverage per-ticker downloads."""
    config = load_daily_update_config(CONFIG_PATH)
    universe_path = tmp_path / "universe.csv"
    raw_path = tmp_path / "raw.parquet"
    clean_path = tmp_path / "clean.parquet"
    pd.DataFrame({"ticker": ["005930", "000660"]}).to_csv(universe_path, index=False)
    values = dict(config.values)
    values["universe_file"] = str(universe_path)
    values["raw_ohlcv_file"] = str(raw_path)
    values["clean_ohlcv_file"] = str(clean_path)
    values["production_mode"] = True
    values["strict_feature_source_check"] = True
    values["min_pykrx_success_ratio"] = 0.95
    test_config = type(config)(config_path=config.config_path, values=values, project_root=PROJECT_ROOT)
    context = build_daily_run_context(test_config, "2026-06-16", False, False, False)

    def partial_downloader(_tickers: list[str], update_date: pd.Timestamp) -> pd.DataFrame:
        rows = valid_ohlcv_rows(update_date.date().isoformat()).query("ticker == '005930'").copy()
        rows.attrs["tickers_requested"] = 2
        rows.attrs["tickers_downloaded"] = 1
        rows.attrs["tickers_failed"] = 1
        rows.attrs["failed_tickers_sample"] = ["000660"]
        rows.attrs["success_ratio"] = 0.5
        rows.attrs["enforce_success_threshold"] = True
        return rows

    result = run_ohlcv_update(test_config, context, False, False, False, downloader=partial_downloader)

    assert result.ohlcv_download_failed is True
    assert result.pykrx_data_unavailable is True
    assert result.pykrx_success_ratio == 0.5
    assert result.pykrx_failed_tickers_sample == ["000660"]
    assert result.raw_rows_added == 0
    assert not raw_path.exists()


def test_high_success_ratio_appends_per_ticker_rows(tmp_path: Path) -> None:
    """Downloads at or above threshold should append successful rows."""
    config = load_daily_update_config(CONFIG_PATH)
    universe_path = tmp_path / "universe.csv"
    raw_path = tmp_path / "raw.parquet"
    clean_path = tmp_path / "clean.parquet"
    pd.DataFrame({"ticker": ["005930", "000660"]}).to_csv(universe_path, index=False)
    values = dict(config.values)
    values["universe_file"] = str(universe_path)
    values["raw_ohlcv_file"] = str(raw_path)
    values["clean_ohlcv_file"] = str(clean_path)
    values["daily_raw_dir"] = str(tmp_path / "daily_raw")
    values["daily_processed_dir"] = str(tmp_path / "daily_processed")
    values["production_mode"] = True
    values["strict_feature_source_check"] = True
    values["min_pykrx_success_ratio"] = 0.95
    test_config = type(config)(config_path=config.config_path, values=values, project_root=PROJECT_ROOT)
    context = build_daily_run_context(test_config, "2026-06-16", False, False, False)

    def complete_downloader(_tickers: list[str], update_date: pd.Timestamp) -> pd.DataFrame:
        rows = valid_ohlcv_rows(update_date.date().isoformat()).copy()
        rows.attrs["tickers_requested"] = 2
        rows.attrs["tickers_downloaded"] = 2
        rows.attrs["tickers_failed"] = 0
        rows.attrs["failed_tickers_sample"] = []
        rows.attrs["success_ratio"] = 1.0
        rows.attrs["enforce_success_threshold"] = True
        return rows

    result = run_ohlcv_update(test_config, context, False, False, False, downloader=complete_downloader)

    assert result.ohlcv_download_failed is False
    assert result.raw_rows_added == 2
    assert result.cleaned_rows_added == 2
    assert result.pykrx_success_ratio == 1.0
    assert raw_path.exists()
    assert len(pd.read_parquet(raw_path)) == 2


def test_normalize_pykrx_ohlcv_missing_columns_raises_unavailable_marker() -> None:
    """Unexpected pykrx payloads should raise a marker consumed by retry logic."""
    with pytest.raises(ValueError, match="pykrx_missing_columns"):
        normalize_pykrx_ohlcv(pd.DataFrame({"unexpected": [1]}), "005930")


def test_run_ohlcv_update_failed_download_falls_back_without_crash(tmp_path: Path) -> None:
    """OHLCV update should not crash when the downloader fails."""
    config = load_daily_update_config(CONFIG_PATH)
    raw_path = tmp_path / "raw.parquet"
    clean_path = tmp_path / "clean.parquet"
    values = dict(config.values)
    values["raw_ohlcv_file"] = str(raw_path)
    values["clean_ohlcv_file"] = str(clean_path)
    test_config = type(config)(config_path=config.config_path, values=values, project_root=PROJECT_ROOT)
    context = build_daily_run_context(test_config, "2026-06-12", False, False, False)

    def failing_downloader(_tickers: list[str], _date: pd.Timestamp) -> pd.DataFrame:
        raise RuntimeError("download failed")

    result = run_ohlcv_update(
        test_config,
        context,
        dry_run=False,
        skip_download=False,
        force=False,
        downloader=failing_downloader,
        timeout_seconds=1,
        max_retries=1,
        retry_sleep_seconds=0,
    )

    assert result.ohlcv_download_failed is True
    assert result.used_existing_data_fallback is True
    assert result.raw_rows_downloaded_or_found == 0
    assert "download failed" in str(result.ohlcv_download_error)


def test_missing_005930_creates_warning(tmp_path: Path) -> None:
    """Daily rows without Samsung should set missing flag and warning."""
    config = load_daily_update_config(CONFIG_PATH)
    raw_path = tmp_path / "raw.parquet"
    clean_path = tmp_path / "clean.parquet"
    values = dict(config.values)
    values["raw_ohlcv_file"] = str(raw_path)
    values["clean_ohlcv_file"] = str(clean_path)
    test_config = type(config)(config_path=config.config_path, values=values, project_root=PROJECT_ROOT)
    context = build_daily_run_context(test_config, "2026-06-12", False, False, False)

    def downloader(_tickers: list[str], update_date: pd.Timestamp) -> pd.DataFrame:
        return valid_ohlcv_rows(update_date.date().isoformat()).query("ticker != '005930'")

    result = run_ohlcv_update(test_config, context, False, False, False, downloader=downloader)

    assert result.missing_005930 is True
    assert "ticker_005930_missing_from_daily_ohlcv" in result.warnings


def sample_macro_rows() -> pd.DataFrame:
    """Create small synthetic macro dataset."""
    return pd.DataFrame(
        {
            "date": pd.to_datetime(["2026-06-10", "2026-06-12", "2026-06-15"]),
            "nasdaq_close": [100.0, 101.0, 999.0],
            "sox_close": [200.0, 201.0, 999.0],
            "sp500_close": [300.0, 301.0, 999.0],
            "vix_close": [20.0, 21.0, 999.0],
            "usdkrw": [1300.0, 1301.0, 999.0],
            "wti_close": [70.0, 71.0, 999.0],
        }
    )


def test_macro_existing_update_date_returns_existing_mode() -> None:
    """Existing update_date should use existing macro row."""
    row, mode, source_date, warnings = build_macro_update_row(sample_macro_rows(), pd.Timestamp("2026-06-12"))

    assert mode == "existing"
    assert source_date == pd.Timestamp("2026-06-12")
    assert warnings == []
    assert row["nasdaq_close"].iloc[0] == 101.0


def test_macro_missing_update_date_forward_fills_from_latest_prior_row() -> None:
    """Missing update_date should forward-fill from latest prior row."""
    row, mode, source_date, warnings = build_macro_update_row(sample_macro_rows(), pd.Timestamp("2026-06-13"))

    assert mode == "forward_fill_prior"
    assert source_date == pd.Timestamp("2026-06-12")
    assert row["date"].iloc[0] == pd.Timestamp("2026-06-13")
    assert row["nasdaq_close"].iloc[0] == 101.0
    assert warnings


def test_macro_forward_fill_does_not_use_future_rows() -> None:
    """Forward-fill must not use rows after update_date."""
    row, _, source_date, _ = build_macro_update_row(sample_macro_rows(), pd.Timestamp("2026-06-13"))

    assert source_date < pd.Timestamp("2026-06-13")
    assert row["nasdaq_close"].iloc[0] != 999.0


def test_macro_raises_if_no_prior_row_exists() -> None:
    """No prior macro row should raise a clear error."""
    try:
        build_macro_update_row(sample_macro_rows(), pd.Timestamp("2026-06-01"))
    except ValueError as exc:
        assert "No prior macro row exists" in str(exc)
    else:
        raise AssertionError("Expected ValueError for missing prior macro row")


def test_macro_dry_run_writes_nothing(tmp_path: Path) -> None:
    """Macro dry-run should not write macro file or snapshot."""
    config = load_daily_update_config(CONFIG_PATH)
    macro_path = tmp_path / "macro.parquet"
    sample_macro_rows().to_parquet(macro_path, index=False)
    values = dict(config.values)
    values["macro_file"] = str(macro_path)
    values["daily_processed_dir"] = str(tmp_path / "daily_processed")
    test_config = type(config)(config_path=config.config_path, values=values, project_root=PROJECT_ROOT)
    context = DailyRunContext(
        run_date="2026-06-13",
        as_of_date="2026-06-13",
        latest_clean_data_date="2026-06-13",
        target_update_date="2026-06-12",
        update_date="2026-06-13",
        prediction_date="2026-06-15",
        dry_run=True,
        skip_download=True,
        force=False,
        warnings=[],
    )

    result = run_macro_update(test_config, context, dry_run=True, force=False)

    assert result.macro_update_mode == "forward_fill_prior"
    assert result.macro_rows_added == 0
    assert result.daily_macro_snapshot_path is None
    assert not (tmp_path / "daily_processed").exists()


def test_macro_force_replace_works() -> None:
    """Force replace should replace update_date row."""
    macro = sample_macro_rows()
    replacement = macro[macro["date"].eq(pd.Timestamp("2026-06-12"))].copy()
    replacement.loc[:, "nasdaq_close"] = 555.0

    updated, rows_added = safe_append_macro(macro, replacement, pd.Timestamp("2026-06-12"), force=True)

    assert rows_added == 1
    assert updated.loc[updated["date"].eq(pd.Timestamp("2026-06-12")), "nasdaq_close"].iloc[0] == 555.0
    assert not updated.duplicated(subset=["date"]).any()


def make_yfinance_frame(date: str = "2026-06-16", close: float = 100.0) -> pd.DataFrame:
    """Create a one-row yfinance-like frame."""
    return pd.DataFrame(
        {
            "Open": [close - 1],
            "High": [close + 1],
            "Low": [close - 2],
            "Close": [close],
            "Volume": [1000],
        },
        index=pd.to_datetime([date]),
    ).rename_axis("Date")


def test_macro_download_target_update_date_and_ticker_mapping() -> None:
    """Production macro download should request required yfinance tickers for target date."""
    calls: list[tuple[str, str, str]] = []

    def fake_downloader(feature: str, ticker: str, target_date: pd.Timestamp) -> pd.DataFrame:
        calls.append((feature, ticker, target_date.date().isoformat()))
        return make_yfinance_frame(target_date.date().isoformat(), close=100.0 + len(calls))

    row = download_required_macro_row(pd.Timestamp("2026-06-16"), downloader=fake_downloader)

    assert MACRO_TICKERS["sox"]["ticker"] == "^SOX"
    assert row["date"].iloc[0] == pd.Timestamp("2026-06-16")
    expected_metadata = {
        f"{kind}_{source}_date"
        for kind in ("actual", "expected")
        for source in MACRO_TICKERS
    }
    assert set(row.columns) == {
        "date", "nasdaq_close", "sp500_close", "vix_close", "wti_close", "usdkrw", "sox_close", *expected_metadata
    }
    assert [call[1] for call in calls] == [info["ticker"] for info in MACRO_TICKERS.values()]
    assert all(call[2] == "2026-06-16" for call in calls)


def test_macro_console_reports_sox_pass_and_fail(capsys) -> None:
    """Operators should see an explicit SOX outcome in both console paths."""
    print_macro_download_success()
    success_output = capsys.readouterr().out

    print_macro_download_failure(MacroDataUnavailableError("SOX", "2026-06-16", "No data returned"))
    failure_output = capsys.readouterr().out

    assert "SOX ......... PASS" in success_output
    assert "SOX ......... FAIL" in failure_output


def test_macro_append_update_duplicate_protection_and_force_replace(tmp_path: Path) -> None:
    """Latest macro append should avoid duplicates and force should replace same date."""
    path = tmp_path / "macro_clean_latest.parquet"
    first = pd.DataFrame(
        {
            "date": pd.to_datetime(["2026-06-16"]),
            "nasdaq_close": [1.0],
            "sp500_close": [2.0],
            "vix_close": [3.0],
            "wti_close": [4.0],
            "usdkrw": [5.0],
        }
    )
    first.to_parquet(path, index=False)
    replacement = first.copy()
    replacement.loc[:, "nasdaq_close"] = 999.0

    unchanged, rows_added, mode = append_latest_macro(path, replacement, pd.Timestamp("2026-06-16"), force=False)
    replaced, replaced_rows, replaced_mode = append_latest_macro(path, replacement, pd.Timestamp("2026-06-16"), force=True)

    assert rows_added == 0
    assert mode == "existing_download_verified"
    assert unchanged.loc[0, "nasdaq_close"] == 1.0
    assert replaced_rows == 1
    assert replaced_mode == "replaced"
    assert replaced.loc[0, "nasdaq_close"] == 999.0
    assert not replaced.duplicated(subset=["date"]).any()


def test_macro_append_existing_date_enriches_missing_sox_contract(tmp_path: Path) -> None:
    """A verified same-date download should fill newly required SOX fields without force."""
    path = tmp_path / "macro_clean_latest.parquet"
    existing = pd.DataFrame(
        {"date": pd.to_datetime(["2026-06-22"]), "nasdaq_close": [100.0]}
    )
    existing.to_parquet(path, index=False)
    downloaded = pd.DataFrame(
        {
            "date": pd.to_datetime(["2026-06-22"]),
            "nasdaq_close": [999.0],
            "sox_close": [200.0],
            "actual_sox_date": pd.to_datetime(["2026-06-22"]),
            "expected_sox_date": pd.to_datetime(["2026-06-22"]),
        }
    )

    updated, rows_added, mode = append_latest_macro(
        path, downloaded, pd.Timestamp("2026-06-22"), force=False
    )

    assert rows_added == 0
    assert mode == "existing_download_enriched"
    assert updated.loc[0, "nasdaq_close"] == 100.0
    assert updated.loc[0, "sox_close"] == 200.0
    assert updated.loc[0, "actual_sox_date"] == pd.Timestamp("2026-06-22")


def test_macro_missing_close_fails() -> None:
    """Missing Close should fail validation."""
    data = make_yfinance_frame().drop(columns=["Close"])

    with pytest.raises(MacroDataUnavailableError, match="Close missing"):
        validate_macro_source_frame(data, "NASDAQ", pd.Timestamp("2026-06-16"))


def test_macro_missing_target_date_fails() -> None:
    """A yfinance row older than the configured tolerance should fail."""
    data = make_yfinance_frame("2026-06-10")

    with pytest.raises(MacroDataUnavailableError, match="older than 5 calendar days"):
        validate_macro_source_frame(data, "NASDAQ", pd.Timestamp("2026-06-16"))


def test_vix_missing_target_close_uses_latest_valid_prior_row() -> None:
    """A null target-date VIX row must not hide a valid prior close."""
    data = pd.DataFrame(
        {"Close": [17.25, float("nan")]},
        index=pd.to_datetime(["2026-06-18", "2026-06-19"]),
    )
    data.index.name = "Date"

    close, actual_date = validate_macro_source_frame(data, "VIX", pd.Timestamp("2026-06-19"))

    assert close == pytest.approx(17.25)
    assert actual_date == pd.Timestamp("2026-06-18")


@pytest.mark.parametrize("source", ["NASDAQ", "SOX", "S&P500", "WTI"])
def test_macro_source_uses_prior_valid_date_within_tolerance(source: str) -> None:
    """All market macro sources share the prior-valid-row policy."""
    data = make_yfinance_frame("2026-06-18", close=100.0)

    close, actual_date = validate_macro_source_frame(data, source, pd.Timestamp("2026-06-19"))

    assert close == pytest.approx(100.0)
    assert actual_date == pd.Timestamp("2026-06-18")


def test_usdkrw_uses_valid_exact_target_date() -> None:
    """An exact valid target row remains preferred."""
    data = make_yfinance_frame("2026-06-19", close=1380.5)

    close, actual_date = validate_macro_source_frame(data, "USD/KRW", pd.Timestamp("2026-06-19"))

    assert close == pytest.approx(1380.5)
    assert actual_date == pd.Timestamp("2026-06-19")


def test_macro_source_fails_when_no_valid_row_is_within_tolerance() -> None:
    """Invalid recent rows cannot make an out-of-tolerance valid row pass."""
    data = pd.DataFrame(
        {"Close": [91.0, float("nan"), -1.0]},
        index=pd.to_datetime(["2026-06-10", "2026-06-18", "2026-06-19"]),
    )
    data.index.name = "Date"

    with pytest.raises(MacroDataUnavailableError, match="No valid Close within 5 calendar days"):
        validate_macro_source_frame(data, "VIX", pd.Timestamp("2026-06-19"))


def test_macro_multiindex_output_selects_latest_valid_row_and_records_diagnostic() -> None:
    """Typical yfinance MultiIndex columns are normalized before validation."""
    columns = pd.MultiIndex.from_tuples([("Close", "^VIX"), ("Open", "^VIX")])
    vix = pd.DataFrame(
        [[17.25, 17.0], [float("nan"), 17.4]],
        index=pd.to_datetime(["2026-06-18", "2026-06-19"]),
        columns=columns,
    )
    vix.index.name = "Date"

    def downloader(source: str, _ticker: str, target_date: pd.Timestamp) -> pd.DataFrame:
        if source == "VIX":
            return vix
        return make_yfinance_frame(target_date.date().isoformat(), close=100.0)

    row = download_required_macro_row(pd.Timestamp("2026-06-19"), downloader=downloader)

    assert row["vix_close"].iloc[0] == pytest.approx(17.25)
    assert row["actual_vix_date"].iloc[0] == pd.Timestamp("2026-06-18")
    assert row["expected_vix_date"].iloc[0] == pd.Timestamp("2026-06-19")
    assert row.attrs["macro_invalid_target_date_rows"] == {"vix": "target-date Close missing"}


@pytest.mark.parametrize(
    ("target_date", "krx_date", "us_date", "holiday_detected"),
    [
        ("2026-06-18", "2026-06-18", "2026-06-18", False),  # both open
        ("2026-06-19", "2026-06-19", "2026-06-18", True),   # US holiday
        ("2026-06-18", "2026-06-18", "2026-06-18", False),  # KRX holiday resolved to prior KRX day
        ("2026-06-18", "2026-06-18", "2026-06-17", True),   # both holidays resolved/aligned
    ],
)
def test_market_aware_feature_source_completeness_passes_valid_market_dates(
    tmp_path: Path,
    target_date: str,
    krx_date: str,
    us_date: str,
    holiday_detected: bool,
) -> None:
    """KRX must match target while US sources may use a recent prior trading day."""
    config = load_daily_update_config(CONFIG_PATH)
    clean_path = tmp_path / "clean.parquet"
    macro_path = tmp_path / "macro.parquet"
    valid_ohlcv_rows(krx_date).to_parquet(clean_path, index=False)
    macro = pd.DataFrame(
        {
            "date": pd.to_datetime([target_date]),
            "nasdaq_close": [100.0],
            "sox_close": [200.0],
            "sp500_close": [100.0],
            "vix_close": [20.0],
            "wti_close": [70.0],
            "usdkrw": [1300.0],
            **{f"actual_{source}_date": pd.to_datetime([us_date]) for source in ("nasdaq", "sox", "sp500", "vix", "wti", "usdkrw")},
        }
    )
    macro.to_parquet(macro_path, index=False)
    values = dict(config.values)
    values["clean_ohlcv_file"] = str(clean_path)
    values["macro_file"] = str(macro_path)
    values["feature_file"] = str(tmp_path / "features_not_built.parquet")
    test_config = type(config)(config_path=config.config_path, values=values, project_root=PROJECT_ROOT)

    result = FeatureSourceCompletenessChecker(test_config, target_date).check()

    assert result["all_available"] is True
    assert result["actual_krx_date"] == krx_date
    assert result["actual_nasdaq_date"] == us_date
    assert result["us_market_holiday_detected"] is holiday_detected


def test_market_aware_feature_source_completeness_fails_missing_nasdaq(tmp_path: Path) -> None:
    """A missing required US close remains a critical failure."""
    config = load_daily_update_config(CONFIG_PATH)
    clean_path = tmp_path / "clean.parquet"
    macro_path = tmp_path / "macro.parquet"
    valid_ohlcv_rows("2026-06-19").to_parquet(clean_path, index=False)
    pd.DataFrame(
        {
            "date": pd.to_datetime(["2026-06-19"]),
            "nasdaq_close": [float("nan")],
            "sox_close": [200.0],
            "sp500_close": [100.0],
            "vix_close": [20.0],
            "wti_close": [70.0],
            "usdkrw": [1300.0],
        }
    ).to_parquet(macro_path, index=False)
    values = dict(config.values)
    values["clean_ohlcv_file"] = str(clean_path)
    values["macro_file"] = str(macro_path)
    values["feature_file"] = str(tmp_path / "features_not_built.parquet")
    test_config = type(config)(config_path=config.config_path, values=values, project_root=PROJECT_ROOT)

    result = FeatureSourceCompletenessChecker(test_config, "2026-06-19").check()

    assert result["all_available"] is False
    assert "nasdaq" in result["failed_sources"]


def test_market_aware_feature_source_completeness_fails_stale_nasdaq(tmp_path: Path) -> None:
    """A US source older than five calendar days must fail."""
    config = load_daily_update_config(CONFIG_PATH)
    clean_path = tmp_path / "clean.parquet"
    macro_path = tmp_path / "macro.parquet"
    valid_ohlcv_rows("2026-06-19").to_parquet(clean_path, index=False)
    macro = pd.DataFrame(
        {
            "date": pd.to_datetime(["2026-06-19"]),
            "nasdaq_close": [100.0],
            "sox_close": [200.0],
            "sp500_close": [100.0],
            "vix_close": [20.0],
            "wti_close": [70.0],
            "usdkrw": [1300.0],
            **{f"actual_{source}_date": pd.to_datetime(["2026-06-18"]) for source in ("sox", "sp500", "vix", "wti", "usdkrw")},
            "actual_nasdaq_date": pd.to_datetime(["2026-06-12"]),
        }
    )
    macro.to_parquet(macro_path, index=False)
    values = dict(config.values)
    values["clean_ohlcv_file"] = str(clean_path)
    values["macro_file"] = str(macro_path)
    values["feature_file"] = str(tmp_path / "features_not_built.parquet")
    test_config = type(config)(config_path=config.config_path, values=values, project_root=PROJECT_ROOT)

    result = FeatureSourceCompletenessChecker(test_config, "2026-06-19").check()

    assert result["all_available"] is False
    assert result["failed_sources"] == ["nasdaq"]


def test_macro_timeout_fails() -> None:
    """A slow macro source should surface as unavailable."""
    release = threading.Event()

    def slow_downloader(_feature: str, _ticker: str, _target_date: pd.Timestamp) -> pd.DataFrame:
        release.wait(1.0)
        return make_yfinance_frame()

    with pytest.raises(MacroDataUnavailableError, match="download timeout"):
        download_required_macro_row(pd.Timestamp("2026-06-16"), downloader=slow_downloader, timeout_seconds=0.01)
    release.set()


def test_run_production_macro_download_writes_latest_dataset(tmp_path: Path) -> None:
    """Production macro downloader should write macro_clean_latest parquet and daily snapshot."""
    config = load_daily_update_config(CONFIG_PATH)
    values = dict(config.values)
    values["daily_processed_dir"] = str(tmp_path / "daily_processed")
    test_config = type(config)(config_path=config.config_path, values=values, project_root=tmp_path)
    context = DailyRunContext(
        run_date="2026-06-17",
        as_of_date="2026-06-17",
        latest_clean_data_date="2026-06-16",
        target_update_date="2026-06-16",
        update_date="2026-06-16",
        prediction_date="2026-06-17",
        dry_run=False,
        skip_download=False,
        force=False,
        warnings=[],
    )

    def fake_downloader(_feature: str, _ticker: str, target_date: pd.Timestamp) -> pd.DataFrame:
        return make_yfinance_frame(target_date.date().isoformat(), close=123.0)

    result = run_production_macro_download(test_config, context, dry_run=False, force=False, downloader=fake_downloader)

    latest_path = tmp_path / "data" / "processed" / "macro" / "macro_clean_latest.parquet"
    assert result.macro_download_method == "yfinance"
    assert result.macro_download_passed is True
    assert result.macro_downloaded_date == "2026-06-16"
    assert result.macro_rows_downloaded == 1
    assert result.actual_source_dates == {source: "2026-06-16" for source in MACRO_TICKERS}
    assert result.expected_source_dates == {source: "2026-06-16" for source in MACRO_TICKERS}
    assert result.macro_invalid_target_date_rows == {}
    assert result.sources_using_prior_trading_day == []
    assert latest_path.exists()
    assert Path(result.daily_macro_snapshot_path).exists()


def test_production_macro_dry_run_reports_latest_macro_path_without_writing(tmp_path: Path) -> None:
    """Dry-run source checks should validate against macro_clean_latest without writing files."""
    config = load_daily_update_config(CONFIG_PATH)
    values = dict(config.values)
    values["daily_processed_dir"] = str(tmp_path / "daily_processed")
    test_config = type(config)(config_path=config.config_path, values=values, project_root=tmp_path)
    context = DailyRunContext(
        run_date="2026-06-23",
        as_of_date="2026-06-23",
        latest_clean_data_date="2026-06-22",
        target_update_date="2026-06-22",
        update_date="2026-06-22",
        prediction_date="2026-06-23",
        dry_run=True,
        skip_download=True,
        force=False,
        warnings=[],
    )

    result = run_production_macro_download(test_config, context, dry_run=True, force=False)

    latest_path = tmp_path / "data" / "processed" / "macro" / "macro_clean_latest.parquet"
    assert result.macro_file_path == str(latest_path)
    assert result.daily_macro_snapshot_path is None
    assert not latest_path.exists()


def test_pipeline_stops_on_macro_download_failure(monkeypatch, tmp_path: Path) -> None:
    """Macro download failure should stop before feature/training/prediction/top10/archive."""
    config = load_daily_update_config(CONFIG_PATH)
    values = dict(config.values)
    values["daily_status_dir"] = str(tmp_path / "status")
    values["daily_report_dir"] = str(tmp_path / "reports")
    values["log_dir"] = str(tmp_path / "logs")
    values["production_mode"] = True
    test_config = type(config)(config_path=config.config_path, values=values, project_root=PROJECT_ROOT)

    monkeypatch.setattr(daily_pipeline, "load_daily_update_config", lambda _path: test_config)
    monkeypatch.setattr(daily_pipeline, "run_ohlcv_update", lambda **_kwargs: _mock_ohlcv_result())

    def fail_macro(**_kwargs):
        raise MacroDataUnavailableError("VIX", "2026-06-16", "No data returned")

    monkeypatch.setattr(daily_pipeline, "run_macro_update", fail_macro)
    monkeypatch.setattr(daily_pipeline, "run_feature_update", lambda **_kwargs: pytest.fail("feature should be skipped"))
    monkeypatch.setattr(daily_pipeline, "run_training_update", lambda **_kwargs: pytest.fail("training should be skipped"))
    monkeypatch.setattr(daily_pipeline, "train_daily_models", lambda **_kwargs: pytest.fail("model should be skipped"))
    monkeypatch.setattr(daily_pipeline, "generate_daily_predictions", lambda **_kwargs: pytest.fail("prediction should be skipped"))
    monkeypatch.setattr(daily_pipeline, "generate_daily_top10_report", lambda **_kwargs: pytest.fail("top10 should be skipped"))
    monkeypatch.setattr(daily_pipeline, "create_daily_archive", lambda **_kwargs: pytest.fail("archive should be skipped"))

    status = daily_pipeline.run_pipeline(
        daily_pipeline.parse_args(["--as-of-date", "2026-06-17", "--skip-download", "--config", str(CONFIG_PATH)])
    )

    assert int(status) == 1
    assert status.pipeline_stop_reason == "Macro Download Failed"
    assert status.macro_download_passed is False
    assert status.macro_download_failed_sources == ["VIX"]
    assert status.top10_generated is False
    assert status.prediction_executed is False


def test_status_includes_macro_fields(monkeypatch, tmp_path: Path) -> None:
    """Normal pipeline status should include macro update fields."""
    as_of_date = "2099-02-03"
    config = load_daily_update_config(CONFIG_PATH)
    values = dict(config.values)
    values["daily_status_dir"] = str(tmp_path / "status")
    values["log_dir"] = str(tmp_path / "logs")
    test_config = type(config)(config_path=config.config_path, values=values, project_root=PROJECT_ROOT)
    status_file = tmp_path / "status" / "daily_update_status_20990203.json"
    monkeypatch.setattr(daily_pipeline, "load_daily_update_config", lambda _path: test_config)
    monkeypatch.setattr(daily_pipeline, "run_ohlcv_update", lambda **_kwargs: _mock_ohlcv_result())
    monkeypatch.setattr(daily_pipeline, "run_macro_update", lambda **_kwargs: _mock_macro_result())
    monkeypatch.setattr(daily_pipeline, "run_feature_update", lambda **_kwargs: _mock_feature_result())
    monkeypatch.setattr(daily_pipeline, "run_training_update", lambda **_kwargs: _mock_training_result())
    monkeypatch.setattr(daily_pipeline, "train_daily_models", lambda **_kwargs: _mock_model_result())
    monkeypatch.setattr(daily_pipeline, "generate_daily_predictions", lambda **_kwargs: _mock_prediction_result())
    monkeypatch.setattr(daily_pipeline, "generate_daily_top10_report", lambda **_kwargs: _mock_report_result())

    daily_pipeline.run_pipeline(
        daily_pipeline.parse_args(["--as-of-date", as_of_date, "--skip-download", "--config", str(CONFIG_PATH)])
    )

    status = json.loads(status_file.read_text(encoding="utf-8"))
    for key in [
        "macro_update_mode",
        "macro_source_date",
        "macro_rows_added",
        "macro_missing_after_update",
        "daily_macro_snapshot_path",
    ]:
        assert key in status
    assert "macro_update_checked" in status["completed_steps"]


def sample_feature_rows(date: str = "2026-06-12") -> pd.DataFrame:
    """Create small synthetic daily feature rows."""
    return pd.DataFrame(
        {
            "date": pd.to_datetime([date, date]),
            "ticker": ["005930", "000660"],
            "return_5d": [0.01, 0.02],
            "return_20d": [0.05, 0.06],
            "momentum_5d": [0.01, 0.02],
            "momentum_20d": [0.05, 0.06],
            "return_5d_rank_pct": [0.5, 1.0],
            "return_rank_pct": [0.5, 1.0],
            "sector": ["Tech", "Tech"],
            "market_type": ["KOSPI", "KOSPI"],
            "market_cap_group": ["Large", "Large"],
            "target_ranking": [0.1, 0.2],
            "feature_a": [1.0, 2.0],
            "sox_return_1d": [0.0, 0.01],
        }
    )


def test_feature_update_preserves_date_and_ticker() -> None:
    """Feature optimization must keep audit keys."""
    optimized = optimize_feature_frame(sample_feature_rows())

    assert "date" in optimized.columns
    assert "ticker" in optimized.columns
    assert optimized["ticker"].tolist() == ["000660", "005930"]


def test_feature_update_excludes_identity_columns() -> None:
    """Identity columns should not be part of optimized model feature output."""
    optimized = optimize_feature_frame(sample_feature_rows())

    assert "sector" not in optimized.columns
    assert "market_type" not in optimized.columns
    assert "market_cap_group" not in optimized.columns


def test_feature_update_excludes_target_columns() -> None:
    """Target columns must not be included in optimized feature output."""
    optimized = optimize_feature_frame(sample_feature_rows())

    assert "target_ranking" not in optimized.columns
    assert all(not column.startswith("target_") for column in optimized.columns)


def test_feature_update_removes_momentum_duplicate_features() -> None:
    """Known duplicate momentum features should be removed."""
    optimized = optimize_feature_frame(sample_feature_rows())

    assert "momentum_5d" not in optimized.columns
    assert "momentum_20d" not in optimized.columns
    assert "return_5d" in optimized.columns
    assert "return_20d" in optimized.columns


def test_feature_update_blocks_entirely_missing_sox() -> None:
    """Production features must not proceed when SOX is null for every ticker."""
    daily = sample_feature_rows()
    daily.loc[:, "sox_return_1d"] = float("nan")

    with pytest.raises(ValueError, match="prior SOX history is missing"):
        validate_required_feature_availability(daily)


def test_feature_safe_append_avoids_duplicate_date_ticker() -> None:
    """Feature safe append should not create duplicate date/ticker rows."""
    existing = optimize_feature_frame(sample_feature_rows())
    daily = optimize_feature_frame(sample_feature_rows())

    combined, rows_added, rows_replaced, mode = safe_append_features(
        existing,
        daily,
        pd.Timestamp("2026-06-12"),
        force=False,
    )

    assert rows_added == 0
    assert rows_replaced == 0
    assert mode == "existing"
    assert not combined.duplicated(subset=["date", "ticker"]).any()


def test_feature_force_replace_replaces_update_date_rows() -> None:
    """Force should replace feature rows for update_date."""
    existing = optimize_feature_frame(sample_feature_rows())
    daily = optimize_feature_frame(sample_feature_rows())
    daily.loc[daily["ticker"].eq("005930"), "feature_a"] = 99.0

    combined, rows_added, rows_replaced, mode = safe_append_features(
        existing,
        daily,
        pd.Timestamp("2026-06-12"),
        force=True,
    )

    assert rows_added == 2
    assert rows_replaced == 2
    assert mode == "replace"
    assert combined.loc[combined["ticker"].eq("005930"), "feature_a"].iloc[0] == 99.0
    assert not combined.duplicated(subset=["date", "ticker"]).any()


def test_feature_dry_run_does_not_write_feature_file_or_snapshot(tmp_path: Path, monkeypatch) -> None:
    """Feature dry-run should not write the feature file or snapshot."""
    config = load_daily_update_config(CONFIG_PATH)
    values = dict(config.values)
    values["feature_file"] = str(tmp_path / "features.parquet")
    values["daily_feature_dir"] = str(tmp_path / "daily_features")
    values["clean_ohlcv_file"] = str(tmp_path / "clean.parquet")
    values["macro_file"] = str(tmp_path / "macro.parquet")
    test_config = type(config)(config_path=config.config_path, values=values, project_root=PROJECT_ROOT)
    context = DailyRunContext(
        run_date="2026-06-12",
        as_of_date="2026-06-12",
        latest_clean_data_date="2026-06-12",
        target_update_date="2026-06-11",
        update_date="2026-06-12",
        prediction_date="2026-06-15",
        dry_run=True,
        skip_download=True,
        force=False,
        warnings=[],
    )

    monkeypatch.setattr("src.pipeline.feature_update.read_ohlcv", lambda _path: valid_ohlcv_rows("2026-06-12"))
    monkeypatch.setattr("src.pipeline.feature_update.read_macro", lambda _path: sample_macro_rows())
    monkeypatch.setattr(
        "src.pipeline.feature_update.read_feature_file",
        lambda _path: pd.DataFrame(columns=["date", "ticker"]),
    )
    monkeypatch.setattr(
        "src.pipeline.feature_update.select_feature_context",
        lambda clean_ohlcv, update_date, lookback_dates=120: clean_ohlcv,
    )
    monkeypatch.setattr(
        "src.pipeline.feature_update.build_daily_features",
        lambda context_ohlcv, macro, update_date: sample_feature_rows(update_date.date().isoformat()),
    )

    result = run_feature_update(test_config, context, dry_run=True, force=False)

    assert result.feature_rows_added == 2
    assert result.daily_feature_snapshot_path is None
    assert not (tmp_path / "features.parquet").exists()
    assert not (tmp_path / "daily_features").exists()


def sample_training_features() -> pd.DataFrame:
    """Create feature rows for daily training update tests."""
    return pd.DataFrame(
        {
            "date": pd.to_datetime(["2026-06-11", "2026-06-11", "2026-06-12", "2026-06-12"]),
            "ticker": ["005930", "000660", "005930", "000660"],
            "feature_a": [1.0, 2.0, 3.0, 4.0],
        }
    )


def sample_target_ohlcv() -> pd.DataFrame:
    """Create OHLCV rows with a next-day target for only one feature date."""
    return pd.DataFrame(
        {
            "date": pd.to_datetime(
                [
                    "2026-06-11",
                    "2026-06-12",
                    "2026-06-11",
                    "2026-06-12",
                ]
            ),
            "ticker": ["005930", "005930", "000660", "000660"],
            "open": [100.0, 102.0, 200.0, 198.0],
            "high": [105.0, 108.0, 205.0, 204.0],
            "low": [95.0, 101.0, 195.0, 197.0],
            "close": [101.0, 106.0, 202.0, 201.0],
            "volume": [1000, 1200, 2000, 2200],
            "trading_value": [101000.0, 127200.0, 404000.0, 442200.0],
        }
    )


def test_training_update_creates_rows_only_when_target_available() -> None:
    """Only feature rows with known next-trading-day targets should enter training."""
    rows = build_target_available_rows(
        sample_training_features(),
        sample_target_ohlcv(),
        pd.Timestamp("2026-06-12"),
    )

    assert len(rows) == 2
    assert set(rows["feature_date"]) == {pd.Timestamp("2026-06-11")}
    assert set(rows["target_date"]) == {pd.Timestamp("2026-06-12")}
    assert {"target_ranking", "target_gap", "target_intraday"} <= set(rows.columns)


def test_training_update_does_not_use_future_unavailable_target() -> None:
    """Rows whose target date is after update_date should not be created."""
    rows = build_target_available_rows(
        sample_training_features(),
        sample_target_ohlcv(),
        pd.Timestamp("2026-06-11"),
    )

    assert rows.empty


def test_training_update_enforces_feature_date_before_target_date() -> None:
    """Generated target rows should satisfy leakage date ordering."""
    rows = build_target_available_rows(
        sample_training_features(),
        sample_target_ohlcv(),
        pd.Timestamp("2026-06-12"),
    )

    assert (rows["feature_date"] < rows["target_date"]).all()
    assert count_leakage_violations(rows) == 0


def test_training_safe_append_avoids_duplicate_date_ticker() -> None:
    """Training append should not duplicate existing target rows."""
    rows = build_target_available_rows(sample_training_features(), sample_target_ohlcv(), pd.Timestamp("2026-06-12"))

    combined, rows_added, rows_replaced, mode = safe_append_training_rows(rows, rows, force=False)

    assert rows_added == 0
    assert rows_replaced == 0
    assert mode == "existing"
    assert not combined.duplicated(subset=["date", "ticker"]).any()
    assert not combined.duplicated(subset=["feature_date", "ticker"]).any()


def test_training_force_replace_works() -> None:
    """Force should replace rows for affected feature dates."""
    existing = build_target_available_rows(sample_training_features(), sample_target_ohlcv(), pd.Timestamp("2026-06-12"))
    replacement = existing.copy()
    replacement.loc[replacement["ticker"].eq("005930"), "feature_a"] = 99.0

    combined, rows_added, rows_replaced, mode = safe_append_training_rows(existing, replacement, force=True)

    assert rows_added == 2
    assert rows_replaced == 2
    assert mode == "replace"
    assert combined.loc[combined["ticker"].eq("005930"), "feature_a"].iloc[0] == 99.0


def test_training_dry_run_writes_nothing(tmp_path: Path, monkeypatch) -> None:
    """Training dry-run should not write dataset or snapshot files."""
    config = load_daily_update_config(CONFIG_PATH)
    values = dict(config.values)
    values["feature_file"] = str(tmp_path / "features.parquet")
    values["clean_ohlcv_file"] = str(tmp_path / "clean.parquet")
    values["training_dataset_file"] = str(tmp_path / "training.parquet")
    values["daily_training_dir"] = str(tmp_path / "daily_training")
    values["rolling_train_days"] = 1
    test_config = type(config)(config_path=config.config_path, values=values, project_root=PROJECT_ROOT)
    context = DailyRunContext(
        run_date="2026-06-12",
        as_of_date="2026-06-12",
        latest_clean_data_date="2026-06-12",
        target_update_date="2026-06-11",
        update_date="2026-06-12",
        prediction_date="2026-06-15",
        dry_run=True,
        skip_download=True,
        force=False,
        warnings=[],
    )
    monkeypatch.setattr("src.pipeline.training_update.read_feature_file", lambda _path: sample_training_features())
    monkeypatch.setattr("src.pipeline.training_update.read_ohlcv", lambda _path: sample_target_ohlcv())
    monkeypatch.setattr(
        "src.pipeline.training_update.read_training_dataset",
        lambda _path: pd.DataFrame(columns=["date", "ticker", "feature_date", "target_date"]),
    )

    result = run_training_update(test_config, context, dry_run=True, force=False)

    assert result.training_rows_added == 2
    assert result.daily_training_snapshot_path is None
    assert not (tmp_path / "training.parquet").exists()
    assert not (tmp_path / "daily_training").exists()


def test_training_update_preserves_prev_close_audit_column() -> None:
    """prev_close should be preserved for pricing audit only."""
    rows = build_target_available_rows(sample_training_features(), sample_target_ohlcv(), pd.Timestamp("2026-06-12"))

    assert "prev_close" in rows.columns
    assert rows.loc[rows["ticker"].eq("005930"), "prev_close"].iloc[0] == 101.0


def test_training_model_feature_exclusion_for_prev_close_and_targets() -> None:
    """prev_close and targets must not be selected as model features."""
    rows = build_target_available_rows(sample_training_features(), sample_target_ohlcv(), pd.Timestamp("2026-06-12"))
    features = get_model_feature_columns(rows)

    assert "prev_close" not in features
    assert "date" not in features
    assert "ticker" not in features
    assert all(not column.startswith("target_") for column in features)


def test_rolling_window_selection_works_after_training_update() -> None:
    """Updated training data should feed the rolling-window selector."""
    dates = pd.bdate_range("2025-01-01", periods=251)
    rows = []
    for feature_date in dates:
        rows.append(
            {
                "date": feature_date + pd.offsets.BDay(1),
                "ticker": "005930",
                "feature_date": feature_date,
                "target_date": feature_date + pd.offsets.BDay(1),
                "prediction_horizon": 1,
                "prev_close": 100.0,
                "feature_a": 1.0,
                "target_ranking": 0.01,
                "target_gap": 0.0,
                "target_intraday": 0.01,
            }
        )
    training = pd.DataFrame(rows)

    train_df, _, _, unique_dates, _ = select_rolling_train_window(
        training,
        dates[-1] + pd.offsets.BDay(1),
        rolling_train_days=250,
    )

    assert len(unique_dates) == 250
    assert train_df["feature_date"].nunique() == 250
    assert train_df["feature_date"].min() == dates[1]


def sample_daily_model_training_dataset(unique_dates: int = 251) -> pd.DataFrame:
    """Create compact numeric training data for daily model tests."""
    dates = pd.bdate_range("2025-01-01", periods=unique_dates)
    rows = []
    for idx, feature_date in enumerate(dates):
        rows.append(
            {
                "date": feature_date + pd.offsets.BDay(1),
                "ticker": "005930",
                "feature_date": feature_date,
                "target_date": feature_date + pd.offsets.BDay(1),
                "prediction_horizon": 1,
                "prev_close": 100.0 + idx,
                "feature_a": float(idx),
                "feature_b": float(idx % 7),
                "target_ranking": 0.001 * idx,
                "target_gap": 0.0001 * idx,
                "target_intraday": 0.0002 * idx,
            }
        )
    return pd.DataFrame(rows)


def test_daily_rolling_retrain_uses_exactly_250_unique_feature_dates(tmp_path: Path) -> None:
    """Daily model training should use a pure 250-date rolling window."""
    config = load_daily_update_config(CONFIG_PATH)
    training_path = tmp_path / "training.parquet"
    sample_daily_model_training_dataset().to_parquet(training_path, index=False)
    values = dict(config.values)
    values["training_dataset_file"] = str(training_path)
    values["daily_model_dir"] = str(tmp_path / "models")
    values["rolling_train_days"] = 250
    values["model_n_estimators_cap"] = 1
    test_config = type(config)(config_path=config.config_path, values=values, project_root=PROJECT_ROOT)

    result = train_daily_models(test_config, pd.Timestamp("2025-12-18"))

    assert result.rolling_train_days == 250
    assert result.train_df["feature_date"].nunique() == 250
    assert result.train_df["feature_date"].max() < pd.Timestamp("2025-12-18")
    assert result.model_paths["ranking_model"].endswith("ranking_model.txt")
    assert Path(result.model_paths["gap_model"]).exists()


def test_daily_model_feature_selection_excludes_forbidden_columns() -> None:
    """Daily model feature selection should exclude audit and target columns."""
    df = sample_daily_model_training_dataset(2)
    features = get_model_feature_columns(df)

    assert features == ["feature_a", "feature_b"]
    assert "prev_close" not in features
    assert "target_gap" not in features
    assert "ticker" not in features


def test_daily_ranking_model_spec_accepts_target_ranking_alias() -> None:
    """Daily ranking training should bridge target_rank_return to target_ranking."""
    config = load_daily_update_config(CONFIG_PATH)
    spec = build_daily_model_spec("ranking_model", sample_daily_model_training_dataset(), config)

    assert spec.target == "target_ranking"


def test_daily_prediction_output_has_required_columns_and_formulas(monkeypatch) -> None:
    """Daily predictions should include required columns and price calculations."""
    input_df = pd.DataFrame(
        {
            "ticker": ["005930", "000660"],
            "ticker_name": ["삼성전자", "SK하이닉스"],
            "prev_close": [100.0, 200.0],
            "feature_a": [1.0, 2.0],
        }
    )
    bundle = SimpleNamespace(
        ranking_model=SimpleNamespace(),
        gap_model=SimpleNamespace(),
        intraday_model=SimpleNamespace(),
    )

    def fake_predict(model, df):
        if model is bundle.ranking_model:
            return pd.Series([0.9, 0.8], index=df.index)
        if model is bundle.gap_model:
            return pd.Series([0.01, -0.02], index=df.index)
        return pd.Series([0.02, 0.03], index=df.index)

    monkeypatch.setattr("src.pipeline.daily_prediction.predict_model", fake_predict)

    predictions = build_daily_prediction_frame(
        input_df,
        bundle,
        prediction_date="2026-06-15",
        train_start_date="2025-06-01",
        train_end_date="2026-06-12",
        rolling_train_days=250,
    )

    required = {
        "prediction_date",
        "ticker",
        "ticker_name",
        "ranking_score",
        "expected_return",
        "pred_gap",
        "pred_intraday",
        "prev_close",
        "pred_open_price",
        "pred_close_price",
        "train_start_date",
        "train_end_date",
        "rolling_train_days",
    }
    assert required <= set(predictions.columns)
    samsung = predictions[predictions["ticker"].eq("005930")].iloc[0]
    assert samsung["expected_return"] == 0.03
    assert samsung["pred_open_price"] == 101.0
    assert samsung["pred_close_price"] == 103.02


def sample_daily_predictions(rows: int = 12) -> pd.DataFrame:
    """Create synthetic daily predictions for report tests."""
    return pd.DataFrame(
        {
            "prediction_date": [pd.Timestamp("2026-06-15")] * rows,
            "ticker": [f"{idx:06d}" for idx in range(1, rows + 1)],
            "ticker_name": [f"Name{idx}" for idx in range(1, rows + 1)],
            "ranking_score": [float(rows - idx) for idx in range(rows)],
            "expected_return": [0.001 * idx for idx in range(rows)],
            "pred_gap": [0.0005 * idx for idx in range(rows)],
            "pred_intraday": [0.0004 * idx for idx in range(rows)],
            "prev_close": [1000.4 + idx for idx in range(rows)],
            "pred_open_price": [1010.6 + idx for idx in range(rows)],
            "pred_close_price": [1020.2 + idx for idx in range(rows)],
        }
    )


def test_daily_report_creates_exactly_10_top10_rows(tmp_path: Path) -> None:
    """Daily report should write exactly ten Top10 rows."""
    config = load_daily_update_config(CONFIG_PATH)
    prediction_path = tmp_path / "predictions.parquet"
    sample_daily_predictions().to_parquet(prediction_path, index=False)
    values = dict(config.values)
    values["daily_report_dir"] = str(tmp_path / "reports")
    test_config = type(config)(config_path=config.config_path, values=values, project_root=PROJECT_ROOT)

    result = generate_daily_top10_report(
        test_config,
        "2026-06-15",
        prediction_parquet=prediction_path,
    )
    top10 = pd.read_csv(result.top10_report_csv, dtype={"ticker": str})

    assert len(top10) == 10
    assert Path(result.top10_report_xlsx).exists()
    assert Path(result.daily_summary_report).exists()


def test_daily_report_ai_score_rank_and_percentile_format() -> None:
    """AI Score, rank, and percentile should be readable and bounded."""
    enriched, _ = enrich_predictions_for_report(sample_daily_predictions())

    assert enriched["AI Score"].between(0, 100).all()
    assert enriched["AI Rank"].iloc[0] == "1 / 12"
    assert enriched["AI Percentile"].iloc[0].startswith("Top ")
    assert enriched["AI Percentile"].iloc[0].endswith("%")


def test_daily_report_price_columns_are_integer_in_csv(tmp_path: Path) -> None:
    """CSV price display columns should be integer KRW values."""
    config = load_daily_update_config(CONFIG_PATH)
    prediction_path = tmp_path / "predictions.parquet"
    sample_daily_predictions().to_parquet(prediction_path, index=False)
    values = dict(config.values)
    values["daily_report_dir"] = str(tmp_path / "reports")
    test_config = type(config)(config_path=config.config_path, values=values, project_root=PROJECT_ROOT)

    result = generate_daily_top10_report(test_config, "2026-06-15", prediction_parquet=prediction_path)
    top10_text = pd.read_csv(result.top10_report_csv, dtype=str)
    for column in ["prev_close", "pred_open_price", "pred_close_price"]:
        assert top10_text[column].str.contains(".", regex=False).sum() == 0
    top10 = pd.read_csv(result.top10_report_csv)
    for column in ["prev_close", "pred_open_price", "pred_close_price"]:
        assert (top10[column] == top10[column].astype(int)).all()


def test_daily_report_has_no_actual_performance_columns(tmp_path: Path) -> None:
    """Live daily report must not include post-close evaluation columns."""
    config = load_daily_update_config(CONFIG_PATH)
    prediction_path = tmp_path / "predictions.parquet"
    sample_daily_predictions().to_parquet(prediction_path, index=False)
    values = dict(config.values)
    values["daily_report_dir"] = str(tmp_path / "reports")
    test_config = type(config)(config_path=config.config_path, values=values, project_root=PROJECT_ROOT)

    result = generate_daily_top10_report(test_config, "2026-06-15", prediction_parquet=prediction_path)
    top10 = pd.read_csv(result.top10_report_csv)

    forbidden = {"Actual Return(%)", "Prediction Error(%)", "Hit", "Prediction Quality"}
    assert forbidden.isdisjoint(set(top10.columns))


def test_daily_report_excel_explanation_sheet_exists(tmp_path: Path) -> None:
    """Excel report should include the Explanation sheet when openpyxl is available."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return
    config = load_daily_update_config(CONFIG_PATH)
    prediction_path = tmp_path / "predictions.parquet"
    sample_daily_predictions().to_parquet(prediction_path, index=False)
    values = dict(config.values)
    values["daily_report_dir"] = str(tmp_path / "reports")
    test_config = type(config)(config_path=config.config_path, values=values, project_root=PROJECT_ROOT)

    result = generate_daily_top10_report(test_config, "2026-06-15", prediction_parquet=prediction_path)
    workbook = load_workbook(result.top10_report_xlsx)

    assert "Explanation" in workbook.sheetnames


def test_daily_archive_created_with_expected_structure(tmp_path: Path) -> None:
    """Successful daily runs should archive Top10, predictions, models, training, status, and metadata."""
    config = load_daily_update_config(CONFIG_PATH)
    test_config = type(config)(config_path=config.config_path, values=dict(config.values), project_root=tmp_path)
    prediction_date = "2026-06-17"
    source_root = tmp_path / "source"
    source_root.mkdir()
    top10_csv = source_root / "top10_20260617.csv"
    top10_xlsx = source_root / "top10_20260617.xlsx"
    summary = source_root / "daily_update_summary_20260617.md"
    prediction_csv = source_root / "predictions_20260617.csv"
    prediction_parquet = source_root / "predictions_20260617.parquet"
    model_dir = source_root / "models"
    model_dir.mkdir()
    for path in [top10_csv, top10_xlsx, summary, prediction_csv]:
        path.write_text("sample", encoding="utf-8")
    pd.DataFrame({"prediction_date": [prediction_date], "ticker": ["005930"]}).to_parquet(prediction_parquet, index=False)
    model_paths = {}
    for model_key in ["ranking_model", "gap_model", "intraday_model"]:
        model_path = model_dir / f"{model_key}.txt"
        model_path.write_text(model_key, encoding="utf-8")
        model_paths[model_key] = str(model_path)
    train_df = pd.DataFrame(
        {
            "feature_date": pd.to_datetime(["2026-06-12"]),
            "target_date": pd.to_datetime(["2026-06-15"]),
            "prediction_horizon": [1],
            "ticker": ["005930"],
            "ticker_name": ["Samsung"],
            "prev_close": [70000.0],
            "feature_a": [0.1],
            "target_ranking": [0.9],
            "target_gap": [0.01],
            "target_intraday": [0.02],
        }
    )
    model_result = SimpleNamespace(
        train_df=train_df,
        feature_columns=["feature_a"],
        train_start_date="2025-01-01",
        train_end_date="2026-06-12",
        rolling_train_days=250,
        model_paths=model_paths,
    )
    prediction_result = SimpleNamespace(
        prediction_output_csv=str(prediction_csv),
        prediction_output_parquet=str(prediction_parquet),
        prediction_rows=1,
    )
    report_result = SimpleNamespace(
        top10_report_csv=str(top10_csv),
        top10_report_xlsx=str(top10_xlsx),
        daily_summary_report=str(summary),
        top10_average_ai_score=88.5,
        top10_average_expected_return=0.0123,
    )
    status = DailyUpdateStatus.create(
        as_of_date=prediction_date,
        dry_run=False,
        skip_download=False,
        force=False,
        rolling_train_days=250,
    )
    status.feature_source_completeness_passed = True
    status.prediction_executed = True
    status.top10_generated = True
    status.target_update_date = "2026-06-16"

    result = create_daily_archive(test_config, prediction_date, model_result, prediction_result, report_result, status)

    archive = Path(result.archive_path)
    latest = tmp_path / "outputs" / "archive" / "latest"
    assert (archive / "top10" / "top10_20260617.csv").exists()
    assert (archive / "top10" / "top10_20260617.xlsx").exists()
    assert (archive / "predictions" / "predictions_20260617.csv").exists()
    assert (archive / "predictions" / "predictions_20260617.parquet").exists()
    assert (archive / "models" / "ranking_model.txt").exists()
    assert (archive / "models" / "gap_model.txt").exists()
    assert (archive / "models" / "intraday_model.txt").exists()
    assert (archive / "training" / "rolling_train_20260617.csv").exists()
    assert (archive / "training" / "rolling_train_20260617.parquet").exists()
    assert (archive / "status" / "daily_update_summary_20260617.md").exists()
    assert (archive / "status" / "daily_update_status_20260617.json").exists()
    assert (archive / "metadata" / "archive_metadata.json").exists()
    assert (archive / "README.txt").exists()
    assert (latest / "top10" / "top10_20260617.csv").exists()
    archived_train = pd.read_parquet(archive / "training" / "rolling_train_20260617.parquet")
    assert archived_train.equals(train_df)
    metadata = json.loads((archive / "metadata" / "archive_metadata.json").read_text(encoding="utf-8"))
    assert metadata["run_date"] == prediction_date
    assert metadata["prediction_date"] == prediction_date
    assert metadata["target_update_date"] == "2026-06-16"
    assert metadata["rolling_train_start_date"] == "2025-01-01"
    assert metadata["rolling_train_end_date"] == "2026-06-12"
    assert metadata["rolling_train_unique_dates"] == 250
    assert metadata["rolling_train_rows"] == 1
    assert metadata["feature_count"] == 1
    assert metadata["model_feature_count"] == 1
    assert metadata["prediction_rows"] == 1
    assert metadata["top10_average_ai_score"] == 88.5
    assert metadata["top10_average_expected_return"] == 0.0123
    assert metadata["python_version"]
    assert metadata["pandas_version"]
    assert metadata["creation_timestamp"]
    assert set(metadata["model_sha256"]) == {"ranking_model", "gap_model", "intraday_model"}
    assert len(metadata["rolling_train_sha256"]) == 64
    assert len(metadata["prediction_sha256"]) == 64
    assert len(metadata["top10_sha256"]) == 64
    assert result.integrity_passed is True
    assert result.sha256_generated is True
    assert Path(result.latest_path).exists()
    assert Path(result.readme_path).exists()
    assert "immutable audit artifacts" in (archive / "README.txt").read_text(encoding="utf-8")


def test_daily_archive_status_copy_and_status_fields(tmp_path: Path) -> None:
    """Final status JSON should be copied into archive and expose archive paths."""
    status = DailyUpdateStatus.create(
        as_of_date="2026-06-17",
        dry_run=False,
        skip_download=False,
        force=False,
        rolling_train_days=250,
    )
    archive_path = tmp_path / "outputs" / "archive" / "20260617"
    status.archive_created = True
    status.archive_path = str(archive_path)
    status.rolling_training_dataset_path_csv = str(archive_path / "training" / "rolling_train_20260617.csv")
    status.rolling_training_dataset_path_parquet = str(archive_path / "training" / "rolling_train_20260617.parquet")
    status.archive_prediction_path = str(archive_path / "predictions")
    status.archive_model_path = str(archive_path / "models")
    status.archive_top10_path = str(archive_path / "top10")
    status.archive_metadata_path = str(archive_path / "metadata" / "archive_metadata.json")
    status.archive_latest_path = str(tmp_path / "outputs" / "archive" / "latest")
    status.archive_integrity_passed = True
    status.archive_sha256_generated = True
    status.archive_readme_path = str(archive_path / "README.txt")
    status_json = tmp_path / "daily_update_status_20260617.json"
    write_status(status, status_json)

    copied = copy_status_into_archive(status_json, status.archive_path, status.as_of_date)

    assert copied.exists()
    saved = json.loads(copied.read_text(encoding="utf-8"))
    assert saved["archive_created"] is True
    assert saved["archive_path"] == status.archive_path
    assert saved["rolling_training_dataset_path_csv"] == status.rolling_training_dataset_path_csv
    assert saved["rolling_training_dataset_path_parquet"] == status.rolling_training_dataset_path_parquet
    assert saved["archive_prediction_path"] == status.archive_prediction_path
    assert saved["archive_model_path"] == status.archive_model_path
    assert saved["archive_top10_path"] == status.archive_top10_path
    assert saved["archive_metadata_path"] == status.archive_metadata_path
    assert saved["archive_latest_path"] == status.archive_latest_path
    assert saved["archive_integrity_passed"] is True
    assert saved["archive_sha256_generated"] is True
    assert saved["archive_readme_path"] == status.archive_readme_path


def test_daily_archive_integrity_fails_when_required_file_missing(tmp_path: Path) -> None:
    """Archive integrity should fail clearly when a required artifact is absent."""
    required = [tmp_path / "exists.txt", tmp_path / "missing.txt"]
    required[0].write_text("ok", encoding="utf-8")

    with pytest.raises(FileNotFoundError, match="Archive integrity failed"):
        verify_archive_integrity(required)


def test_daily_runner_bat_files_exist() -> None:
    """One-click Windows runner files should exist."""
    assert (PROJECT_ROOT / "run_daily_update.bat").exists()
    assert (PROJECT_ROOT / "run_daily_update_dry_run.bat").exists()


def test_daily_runner_bat_files_report_exit_codes() -> None:
    """Windows runners should print final exit-code meaning before pausing."""
    for filename in [
        "run_daily_update.bat",
        "run_daily_update_dry_run.bat",
        "run_daily_update_no_download.bat",
        "run_daily_update_no_login.bat",
    ]:
        text = (PROJECT_ROOT / filename).read_text(encoding="utf-8")
        assert "Pipeline finished." in text
        assert "Exit code: %EXIT_CODE%" in text
        assert "Meaning: Failed - production feature source completeness check failed." in text
        assert "0 = success" in text
        assert "1 = failed" in text
        assert "130 = user interrupted" in text
